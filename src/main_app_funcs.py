#!/usr/bin/env python
# -*- coding: utf-8 -*- #
import shutil
import sqlite3
import tkinter
import os
from tkinter import *
from tkinter import ttk, filedialog, messagebox
from os.path import abspath

from openpyxl.reader.excel import load_workbook
from pandas import DataFrame
from PIL import Image, ImageTk
from tkinter.ttk import Frame
from idlelib.tooltip import Hovertip
from typing import Callable, Any, List, Literal

from src.global_values import user_data
from src.data import mold_statuses_list, part_statuses_list, columns_warehouse_table, \
    columns_sizes_warehouse_table, columns_bom_parts_table, columns_sizes_bom_parts_table, columns_purchased_parts, \
    columns_sizes_purchased_parts_table, columns_molds_table, columns_sizes_molds_table
from src.data import info_messages, error_messages, columns_molds_moving_history_table, \
    columns_sizes_moving_history_table
from src.utils.excel.xls_tables import get_new_bom_from_excel_file, export_excel_table
from src.utils.gui.bom_edition_funcs import EditedBOM
from src.utils.gui.mold_status_changing_funcs import QRWindow
from src.utils.gui.necessary_spare_parts_report_funcs import MinPartsReport, get_mold_names_list
from src.utils.gui.purchased_parts_funcs import upload_purchased_parts, CustomsReport, PurchasedPart
from src.utils.gui.reference_information_funcs import ReferenceInfo
from src.utils.gui.spare_parts_searching_funcs import Searcher
from src.utils.gui.user_authorization_funcs import change_user
from src.utils.gui.users_data_edition import UsersData
from src.utils.logger.logs import get_info_log, get_warning_log
from src.utils.sql_database import table_funcs, new_tables
from src.utils.gui.mold_data_edition_funcs import EditedMold
from src.utils.gui.stock_changing_funcs import Stock
from src.utils.gui.attached_files_review_funcs import Attachment
from src.utils.sql_database.table_funcs import DataBase, TableInDb


def validate_new_bom(mold_number: str, column_names: tuple, rows_data: list, hot_runner: bool = None) -> (bool, str):
    """
    Функция проверки наличия номера пресс-формы в общем перечне всех п/ф, проверки на наличие уже созданного ранее BOM
    с таким номером п/ф, а также проверки на соответствие названий столбцов таблицы из нового BOM
    :param rows_data: Все строки с информацией из Excel файла загружаемого BOM
    :param mold_number: Номер пресс-формы
    :param column_names: Наименования столбцов нового BOM
    :param hot_runner: Булево значение, которое характеризует какой тип BOM был выбран (Пресс-форма или горячий канал)
    :return: True - если номер п/ф существует в перечне и такой BOM не создавался ранее
    """
    define_table_name: Callable = lambda: f'BOM_HOT_RUNNER_{mold_number}' if hot_runner else f'BOM_{mold_number}'
    # Выгрузка информации из базы данных
    molds_data = table_funcs.TableInDb('All_molds_data', 'Database')
    molds_table = molds_data.get_table(type_returned_data='tuple')
    # Поиск соответствия по номеру пресс-формы в общем перечне
    for mold_info in molds_table:
        if mold_info[0] == mold_number:
            db = table_funcs.DataBase('Database')
            tables = db.get_all_tables()
            # Проверка базы данных на наличие схожей таблицы по названию
            new_table = define_table_name()
            for table in tables:
                if table[0] == new_table:
                    return False, 'Таблица с таким названием уже имеется в базе данных'
            # Проверка названий столбцов нового BOM на корректность
            for num, name in enumerate(columns_bom_parts_table):
                try:
                    if name != column_names[num]:
                        return False, (f'Столбец: {name} отсутствует, либо расположен в несоответствии с шаблоном.'
                                       f'\nСравните вашу таблицу с шаблоном.')
                except IndexError:
                    return False, (f'Столбец: {name} отсутствует, либо расположен в несоответствии с шаблоном.'
                                   f'\nСравните вашу таблицу с шаблоном.')

            part_nums = []
            for count, row in enumerate(rows_data):
                # Проверка на наличие номера запчасти в каждой строке и на то, чтобы этот номер не повторялся в BOM
                part_num = row[0]
                pcs_in_mold = row[2]
                new_parts_in_stock = row[8]
                old_parts_in_stock = row[9]
                min_percent = row[11]
                # Проверка на наличие номера запчасти в каждой строке
                if not part_num:
                    return (column_names, rows_data,
                            False, 'BOM не может быть загружен, так как имеется строка без номера запчасти')
                # Проверка на то, чтобы номер элемента не повторялся в BOM
                elif part_num in part_nums:
                    return (column_names, rows_data, False,
                            f'BOM не может быть загружен, так как номер запчасти: {part_num} '
                            f'дублируется в строке {count + 1}')
                # Проверки на соответствие значений числовому типу данных
                elif count != 0 and not isinstance(pcs_in_mold, int):
                    return (column_names, rows_data, False,
                            f'BOM не может быть загружен, так как значение "Кол-во в пресс-форме, шт": {pcs_in_mold}'
                            f' должно быть числом в строке {count + 1}')
                elif count != 0 and not isinstance(new_parts_in_stock, int):
                    return (column_names, rows_data, False,
                            f'BOM не может быть загружен, так как значение "Кол-во в пресс-форме, шт": {new_parts_in_stock}'
                            f' должно быть числом в строке {count + 1}')
                elif count != 0 and not isinstance(old_parts_in_stock, int):
                    return (column_names, rows_data, False,
                            f'BOM не может быть загружен, так как значение "Кол-во в пресс-форме, шт": {old_parts_in_stock}'
                            f' должно быть числом в строке {count + 1}')
                elif count != 0 and not isinstance(min_percent, int):
                    return (column_names, rows_data, False,
                            f'BOM не может быть загружен, так как значение "Кол-во в пресс-форме, шт": {min_percent}'
                            f' должно быть числом в строке {count + 1}')
                part_nums.append(part_num)

            return True

    return False, (f'BOM не может быть загружен, так как с именем: {mold_number} не найдено ни одной пресс-формы '
                   f'из общего перечня.'
                   f'\n\nЕсли вы загружаете спецификацию горячего канала, убедитесь, '
                   f'что файл называется номером ПРЕСС-ФОРМЫ, а не горячего канала.')


def create_menu_widgets(window: tkinter.Tk, application):
    """
    Функция для создания виджетов с выпадающими списками команд, которые находятся под строкой заголовка наверху
    :param window: Окно приложения
    :param application: Экземпляр класса приложения
    """
    menu = Menu(window)
    window.config(menu=menu)
    file_menu = Menu(menu, tearoff=0)
    file_menu.add_command(label='Сменить пользователя', command=lambda: change_user(window))
    if user_data.get('user_name') == 'admin':
        file_menu.add_command(label='Редактировать пользователей', command=lambda: UsersData().render_widgets())
    menu.add_cascade(label='Файл', menu=file_menu, )
    window.config(menu=menu)
    operations_menu = Menu(menu, tearoff=0)
    operations_menu.add_command(label='Перечень п/ф', command=application.get_molds_data)
    operations_menu.add_command(label='Дефектация', command=lambda: application.render_typical_additional_window(
        called_class=MinPartsReport,
        window_name='Min Parts Searching')),
    operations_menu.add_command(label='Журнал склада',
                                command=lambda: application.open_warehouse_history_window(consumption=True))
    operations_menu.add_command(label='История перемещений п/ф', command=application.open_mold_scanning_window)
    operations_menu.add_command(label='Поиск запчастей', command=lambda: application.render_typical_additional_window(
        called_class=Searcher,
        window_name='Spare Parts Searching'))
    menu.add_cascade(label='Открыть', menu=operations_menu)
    help_menu = Menu(menu, tearoff=0)
    help_menu.add_command(label='Помощь', command=lambda: application.render_typical_additional_window(
        called_class=ReferenceInfo,
        window_name='Reference Information'))
    help_menu.add_command(label='О программе', command=lambda: application.render_typical_additional_window(
        called_class=lambda: ReferenceInfo(app_info=True),
        window_name='App Info'))
    menu.add_cascade(label='Справка', menu=help_menu)


def render_window_for_selection_bom_type(window_title: str, callback_func: Callable):
    """
    Рендер окна для выбора типа BOM для его загрузки в систему приложения, либо для скачиваемого пустого шаблона BOM
    (пресс-форма или горячий канал)
    :param window_title: Имя доп. окна
    :param callback_func: Вызываемая функция при нажатии на кнопку
    """
    window = tkinter.Toplevel()
    window.title(window_title)
    window.resizable(False, False)
    window.focus_set()
    window.grab_set()
    (ttk.Label(window, text='Выберите тип BOM', font=('Times', '11', 'normal'))
     .pack(side=TOP, pady=5))
    ttk.Button(
        window, text='Пресс-форма', style='Regular.TButton',
        width=20,
        command=lambda: callback_func(window)
    ).pack(side=LEFT, padx=5, pady=5)
    ttk.Button(
        window, text='Горячий канал', style='Regular.TButton',
        width=20,
        command=lambda: callback_func(window, hot_runner_bom=True)
    ).pack(side=LEFT, padx=5, pady=5)
    window.mainloop()


def check_value_type(value: Any) -> bool:
    """
    Функция возвращает булево значения соответствующее типу передаваемой переменной
    :param value: Переменная тип, который необходимо определить (строка или цифра)
    :return: Булево значение
    """
    try:
        int(value)
    except ValueError:
        return False
    else:
        return True


def fill_bom_parameters():
    # Выгрузка данных из иксель файла
    work_book = load_workbook(os.path.join('parameters.xls'))
    parameters = []
    try:
        work_sheet = work_book['data']
    except KeyError:
        messagebox.showerror('Уведомление об ошибке',
                             'Название листа не соответствует. '
                             'Убедитесь, что вы используете правильный файл шаблона')
    else:
        for count, row in enumerate(work_sheet.values):
            # Проверка на наличие номера запчасти в каждой строке и на то, чтобы этот номер не повторялся в BOM
            parameters.append(row)
    # Сортировка бд на поиск схожих значений с последующей замены данных
    bom_table_names = get_mold_names_list()
    for bom in bom_table_names:
        db = TableInDb(bom, 'Database')
        table_data = db.get_table(type_returned_data='dict')
        for row in table_data:
            for param in parameters:
                pass


class App(Frame):
    """
    Класс представляет набор функций для создания графического интерфейса основного окна приложения с помощью
    библиотеки Tkinter. А именно отрисовка / вывод различных изображений, таблиц, кнопок, полей ввода и т.д. 
    Помимо рендера виджетов, класс включает в себя функции для обработки отображаемой информации. В частности это 
    взаимодействие с таблицами базы данных.
    """

    def __init__(self):
        """
        Инициация переменных класса
        self.tree: Контейнер для вывода информации в табличном виде
        self.frame_header: Контейнер для отображения шапки приложения
        self.frame_main_widgets: Контейнер для отображения кнопок и полей ввода ниже шапки приложения
        self.frame_body: Контейнер для отображения основной информации
        self.tracked_variable: Переменная для отслеживания записываемых данных в поле ввода
        self.hot_runner_bom: Булево значение, которое становится True когда выбирается пользователем таблица
        горячего канала
        self.sorted_bom_tuple: Словарь, содержащий данные (BOM выбранной пресс-формы) в кортежах для вывода
        отсортированной информации по разным признакам
        self.all_molds_table_dict: Словарь, содержащий данные (общий перечень пресс-форм) в словарях
        self.mold_number: Текстовое значение параметра "MOLD_NUMBER" из таблицы базы данных. Принимает значение
        выбранной пресс-формы пользователем в данный момент
        self.mold_number_entry_field: Поле ввода для открытия BOM по номеру пресс-формы
        self.current_table: Массив состоящий из кортежей с данными для вывода таблицы в окне приложения
        self.event: Обрабатываемое событие
        self.sort_status: Значение характеризующее выбранный параметр сортировки для вывода табличных данных
        self.sorted_molds_data_tuple: Словарь, содержащий данные (общий перечень пресс-форм) в кортежах для вывода
        отсортированной информации по разным признакам
        self.sorted_molds_data_dict: Словарь, содержащий данные (общий перечень пресс-форм) в словарях для вывода
        отсортированной информации по разным признакам
        self.molds_list_data: Словарь, содержащий данные (общий перечень пресс-форм) в кортежах
        """
        super().__init__()
        self.scroll_y = None
        self.bom_data = None
        self.bom_table_dict = None
        self.sorted_bom_dict = None
        self.hot_runner_bom = None
        self.sorted_bom_tuple = None
        self.all_molds_table_dict = None
        self.mold_number = None
        self.current_table = None
        self.mold_number_entry_field = None
        self.event = None
        self.user_name = user_data.get('user_name')
        # Объявление переменных изображений
        self.image_logo = Image.open(os.path.abspath(os.path.join('pics', 'company_logo.png'))) \
            .resize((150, 70))
        self.image_logo_pil = ImageTk.PhotoImage(self.image_logo)
        self.image_body = Image.open(os.path.abspath(os.path.join('pics', 'main__picture.png'))) \
            .resize((800, 500))
        self.image_body_pil = ImageTk.PhotoImage(self.image_body)
        self.main_menu_icon = Image.open(os.path.abspath(os.path.join('pics', 'main_menu.png'))) \
            .resize((20, 20))
        self.main_menu_icon_pil = ImageTk.PhotoImage(self.main_menu_icon)
        self.image_mold = Image.open(os.path.abspath(os.path.join('pics', 'mold.png'))) \
            .resize((144, 120))
        self.image_mold_pil = ImageTk.PhotoImage(self.image_mold)
        self.image_spare_part = Image.open(os.path.abspath(os.path.join('pics', 'spare_parts.png'))) \
            .resize((144, 120))
        self.image_spare_part_pil = ImageTk.PhotoImage(self.image_spare_part)
        self.image_scanner = Image.open(os.path.abspath(os.path.join('pics', 'scanner.png'))) \
            .resize((144, 120))
        self.image_scanner_pil = ImageTk.PhotoImage(self.image_scanner)
        self.image_rack = Image.open(os.path.abspath(os.path.join('pics', 'rack.png'))) \
            .resize((144, 120))
        self.image_rack_pil = ImageTk.PhotoImage(self.image_rack)
        self.back_icon = Image.open(os.path.abspath(os.path.join('pics', 'back.png'))) \
            .resize((20, 20))
        self.back_icon_pil = ImageTk.PhotoImage(self.back_icon)
        self.plus_icon = Image.open(os.path.abspath(os.path.join('pics', 'plus.png'))) \
            .resize((20, 20))
        self.plus_icon_pil = ImageTk.PhotoImage(self.plus_icon)
        self.delete_icon = Image.open(os.path.abspath(os.path.join('pics', 'delete.png'))) \
            .resize((20, 20))
        self.delete_icon_pil = ImageTk.PhotoImage(self.delete_icon)
        self.edit_icon = Image.open(os.path.abspath(os.path.join('pics', 'edit.png'))) \
            .resize((20, 20))
        self.edit_icon_pil = ImageTk.PhotoImage(self.edit_icon)
        self.excel_icon = Image.open(os.path.abspath(os.path.join('pics', 'excel.png'))) \
            .resize((20, 20))
        self.excel_icon_pil = ImageTk.PhotoImage(self.excel_icon)
        self.info_icon = Image.open(os.path.abspath(os.path.join('pics', 'info.png'))) \
            .resize((20, 20))
        self.info_icon_pil = ImageTk.PhotoImage(self.info_icon)
        self.help_icon = Image.open(os.path.abspath(os.path.join('pics', 'help.png'))) \
            .resize((20, 20))
        self.help_icon_pil = ImageTk.PhotoImage(self.help_icon)
        self.new_attachment_icon = Image.open(os.path.abspath(os.path.join('pics', 'new_attachment.png'))) \
            .resize((20, 20))
        self.new_attachment_icon_pil = ImageTk.PhotoImage(self.new_attachment_icon)
        self.attachments_icon = Image.open(os.path.abspath(os.path.join('pics', 'attachments.png'))) \
            .resize((20, 20))
        self.attachments_icon_pil = ImageTk.PhotoImage(self.attachments_icon)
        self.loupe_icon = Image.open(os.path.abspath(os.path.join('pics', 'loupe.png'))) \
            .resize((20, 20))
        self.loupe_icon_pil = ImageTk.PhotoImage(self.loupe_icon)
        self.open_attachments_button = None
        # Объявление переменных будущих контейнеров для хранения виджетов
        self.tree = ttk.Treeview()
        self.frame_header = None
        self.frame_toolbar = None
        self.frame_toolbar_attachments = None
        self.frame_main_widgets = None
        self.frame_body = None
        self.frame_bottom = None
        # Объявление переменных для сортировки данных
        self.sort_status = None
        self.sorted_molds_data_tuple = None
        self.sorted_molds_data_dict = None
        self.molds_list_data = None

        self.init_gui()

    def init_gui(self):
        """
        Инициация окна приложения, а также основных контейнеров
        """
        self.master.title(f'MoldShop Management ({user_data.get("user_name")})')
        self.pack(fill=BOTH)
        self.frame_header = Frame(self, relief=RAISED)
        self.frame_header.pack(fill=BOTH, expand=True)
        self.frame_main_widgets = Frame(self)
        self.frame_main_widgets.pack(anchor=N)
        self.frame_body = Frame(self)
        self.frame_body.pack()
        self.frame_bottom = Frame(self)
        self.frame_bottom.pack()

    def upload_new_bom(self, window: tkinter, hot_runner_bom: bool = None):
        """
        Функция загрузки спецификации пресс-формы (BOM) из Иксель файла формата xlsx в таблицу базы данных
        :param window: дополнительное окно приложения
        :param hot_runner_bom: Булево значение, которое характеризует какой тип BOM был выбран
        (Пресс-форма или горячий канал)
        """
        # Открытие диалогового окна для выбора файла пользователем с локальной директории компьютера,
        # с дальнейшим извлечением пути к выбранному файлу в виде строки
        if user_data.get('molds_and_boms_data_changing') == 'True':
            try:
                window.quit()
                window.destroy()
                file_path = filedialog.askopenfile(
                    filetypes=(('XLSX files', '*.xlsx'),)
                ).name
            except AttributeError:
                pass
            else:
                # Получение информации из Иксель файла типа xlsx
                define_sheet_name: Callable = lambda: 'HOT RUNNER' if hot_runner_bom else 'MOLD'
                try:
                    column_names, rows_data, status, error_message = get_new_bom_from_excel_file(
                        file_path=file_path,
                        work_sheet_name=define_sheet_name())
                    if not status:
                        messagebox.showerror(title=error_messages.get('not_downloaded_bom').get('message_name'),
                                             message=error_message)
                        get_warning_log(user=user_data.get('user_name'), message='New BOM wasnt uploaded',
                                        func_name=self.upload_new_bom.__name__, func_path=abspath(__file__))
                        return
                except TypeError:
                    get_warning_log(user=user_data.get('user_name'), message='New BOM wasnt uploaded',
                                    func_name=self.upload_new_bom.__name__, func_path=abspath(__file__))
                else:
                    file_path = file_path.split('/')
                    mold_number = file_path[-1].replace('.xlsx', '')
                    # Поиск соответствия по номеру пресс-формы в общем перечне
                    status, message = validate_new_bom(mold_number=mold_number, hot_runner=hot_runner_bom,
                                                       column_names=column_names, rows_data=rows_data)
                    if status:
                        # Сохранение информации в базе данных
                        new_tables.create_bom_for_new_mold(mold_number=mold_number, rows_data=rows_data,
                                                           hot_runner=hot_runner_bom)
                        # Рендер окна приложения с новой загруженной информацией
                        info_message = info_messages.get('downloaded_bom').get('message_body')
                        self.mold_number = mold_number
                        self.open_bom(mold_number=mold_number)
                        messagebox.showinfo(title=info_messages.get('downloaded_bom').get('message_name'),
                                            message=info_message.format(mold_number=mold_number))

                        get_info_log(user=self.user_name, message='New BOM was uploaded',
                                     func_name=self.upload_new_bom.__name__,
                                     func_path=abspath(__file__))
                    else:
                        messagebox.showerror(title=error_messages.get('not_downloaded_bom').get('message_name'),
                                             message=message)
                        get_warning_log(user=self.user_name, message='New BOM was NOT uploaded',
                                        func_name=self.upload_new_bom.__name__,
                                        func_path=abspath(__file__))
        else:
            messagebox.showerror(error_messages.get('access_denied').get('message_name'),
                                 error_messages.get('access_denied').get('message_body'))

    def upload_attachment(self, bom_part: bool = None):
        """
        Функция загрузки вложения (какого либо файла) в директорию приложения для дальнейшего использования
        :param bom_part: Булево значение, которое принимает True когда открыт какой-либо BOM
        """
        # Открытие диалогового окна для выбора файла пользователем с локальной директории компьютера,
        # с дальнейшим извлечением пути к выбранному файлу в виде строки
        part_number = None
        if user_data.get('attachments_changing') == 'True':
            if bom_part:
                part_number = self.get_selected_row_data()[0]
            if not self.mold_number:
                self.mold_number = self.get_selected_row_data()[0]
            if self.mold_number:
                try:
                    file_path = filedialog.askopenfile().name
                except AttributeError:
                    pass
                else:
                    # Получение имени файла
                    file_path_list = file_path.split('/')
                    file_name = file_path_list[-1]
                    # Создание нужной директории где будет размещён файл если она ещё не была создана
                    define_folder: Callable = lambda: os.path.join('savings', 'attachments', self.mold_number) \
                        if not part_number \
                        else (
                        os.path.join('savings', 'attachments', self.mold_number, 'hot_runner_parts', part_number)
                        if self.hot_runner_bom else os.path.join('savings', 'attachments', self.mold_number,
                                                                 'mold_parts',
                                                                 part_number))
                    try:
                        os.mkdir(os.path.join('savings', 'attachments', self.mold_number))
                    except FileExistsError:
                        pass
                    try:
                        os.mkdir(os.path.join('savings', 'attachments', self.mold_number, 'hot_runner_parts'))
                        os.mkdir(os.path.join('savings', 'attachments', self.mold_number, 'mold_parts'))
                    except FileExistsError:
                        pass
                    if bom_part:
                        try:
                            os.mkdir(define_folder())
                        except FileExistsError:
                            pass

                    # Копирование и вставка файла в директорию приложения
                    define_path: Callable = lambda: os.path.join('savings', 'attachments', self.mold_number, file_name) \
                        if not bom_part \
                        else (
                        os.path.join('savings', 'attachments', self.mold_number, 'hot_runner_parts',
                                     part_number, file_name)
                        if self.hot_runner_bom else os.path.join('savings', 'attachments', self.mold_number,
                                                                 'mold_parts', part_number, file_name))
                    try:
                        shutil.copy2(file_path, os.path.abspath(define_path()))
                    except IOError:
                        messagebox.showerror(title='Ошибка',
                                             message='Файл не удалось прикрепить. Обратитесь к администратору.')
                        get_warning_log(user=user_data.get('user_name'), message='File wasnt attached because of '
                                                                                 'IOError',
                                        func_name=self.upload_attachment.__name__, func_path=abspath(__file__))
                    else:
                        messagebox.showinfo(title='Уведомление', message='Файл успешно прикреплён')
                        self.render_typical_additional_window(called_class=lambda: Attachment(
                            mold_number=self.mold_number, part_number=part_number, hot_runner=self.hot_runner_bom),
                                                              window_name='Attachments')
                        get_info_log(user=self.user_name, message='Document was attached',
                                     func_name=self.upload_attachment.__name__,
                                     func_path=abspath(__file__))
            else:
                messagebox.showerror(title='Ошибка',
                                     message='Чтобы прикрепить файл выберите элемент из таблицы')
                get_warning_log(user=self.user_name, message='Document was NOT attached',
                                func_name=self.upload_attachment.__name__,
                                func_path=abspath(__file__))
        else:
            messagebox.showerror(error_messages.get('access_denied').get('message_name'),
                                 error_messages.get('access_denied').get('message_body'))

    def save_bom_parts_template(self, window: tkinter, hot_runner_bom: bool = None):
        """
        Функция сохранения шаблона формирования BOM в Иксель файл формата xlsx
        :param window: Открытое дополнительное окно приложения для выбора типа пресс-формы
        :param hot_runner_bom: Булево значение, которое становится True когда выбирается пользователем таблица
        горячего канала
        """
        # Открытие диалогового окна для сохранения файла пользователем в локальной директории компьютера,
        # с дальнейшим извлечением пути к выбранному файлу в виде строки
        import openpyxl
        from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
        from openpyxl.utils import get_column_letter

        try:
            window.quit()
            window.destroy()
            file_path = filedialog.asksaveasfilename(
                filetypes=(('XLSX files', '*.xlsx'),)
            )
        except AttributeError:
            pass
        else:
            if file_path:
                file_path = file_path.replace('.xlsx', '')
                # Формирование Иксель файла типа xlsx
                saved_table = {}
                for num, column_name in enumerate(columns_bom_parts_table):
                    saved_table[column_name] = tuple()
                try:
                    define_sheet_name: Callable = lambda: 'HOT RUNNER' if hot_runner_bom else 'MOLD'
                    df = DataFrame(saved_table)
                    df.to_excel(excel_writer=f'{file_path}.xlsx', sheet_name=define_sheet_name(), index=False)
                    wb = openpyxl.load_workbook(f'{file_path}.xlsx')
                    ws = wb[define_sheet_name()]
                    dim_holder = DimensionHolder(worksheet=ws)

                    for col in range(ws.min_column, ws.max_column + 1):
                        dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=20)

                    ws.column_dimensions = dim_holder
                    wb.save(f'{file_path}.xlsx')
                except Exception:
                    messagebox.showerror(title='Уведомление об ошибке',
                                         message='Ошибка в записи файла.\nПовторите ещё раз, либо обратитесь к '
                                                 'администратору')
                    get_warning_log(user=self.user_name, message='Template was not saved in Excel file',
                                    func_name=self.save_bom_parts_template.__name__, func_path=abspath(__file__))
                else:
                    messagebox.showinfo(title='Уведомление', message='BOM шаблон успешно сохранен на Ваш компьютер')
                    get_info_log(user=self.user_name, message='BOM template was saved',
                                 func_name=self.save_bom_parts_template.__name__,
                                 func_path=abspath(__file__))

    def render_toolbar(self, back_func: Callable, add_row_func: Callable = None, edit_row_func: Callable = None,
                       delete_row_func: Callable = None, new_attachment_func: Callable = None,
                       looking_attachments_func: Callable = None):
        """
        Рендер виджетов панели инструментов
        :param looking_attachments_func: Функция просмотра вложенных файлов
        :param new_attachment_func: Функция прикрепления файла
        :param back_func: Вызываемая функция при нажатии кнопки "Назад" 
        :param add_row_func: Функция добавления новой строки в открытой таблице
        :param edit_row_func: Функция редактирования выбранной строки в открытой таблице
        :param delete_row_func: Функция удаления выбранной строки в открытой таблице
        """
        # Объявление контейнеров
        self.frame_toolbar = ttk.Frame(self)
        self.frame_toolbar.pack(fill=BOTH, expand=True)
        frame_toolbar_menu = Frame(self.frame_toolbar, relief=RIDGE)
        frame_toolbar_menu.pack(side=LEFT, padx=0)
        frame_toolbar_table = Frame(self.frame_toolbar, relief=RIDGE)
        frame_toolbar_table.pack(side=LEFT, padx=0)
        self.frame_toolbar_attachments = Frame(self.frame_toolbar, relief=RIDGE)
        self.frame_toolbar_attachments.pack(side=LEFT, padx=0)
        frame_toolbar_searching = Frame(self.frame_toolbar, relief=RIDGE)
        frame_toolbar_searching.pack(side=LEFT, padx=0)
        frame_toolbar_picture = Frame(self.frame_toolbar, relief=RIDGE)
        frame_toolbar_picture.pack(side=LEFT, padx=0)
        frame_toolbar_info = Frame(self.frame_toolbar, relief=RIDGE)
        frame_toolbar_info.pack(side=RIGHT, padx=0)
        # Объявление виджетов и их рендер
        ttk.Label(frame_toolbar_menu, text=f'Меню', anchor=NW, style='Toolbar.TLabel').pack(expand=True, padx=2, pady=2)
        menu_button = ttk.Button(frame_toolbar_menu, image=self.main_menu_icon_pil,
                                 command=self.open_main_menu)
        menu_button.pack(side=LEFT, padx=3, pady=4)
        Hovertip(anchor_widget=menu_button, text='Главное меню', hover_delay=400)
        back_button = ttk.Button(frame_toolbar_menu, image=self.back_icon_pil,
                                 command=back_func)
        back_button.pack(side=LEFT, padx=3, pady=4)
        Hovertip(anchor_widget=back_button, text='Назад', hover_delay=400)

        ttk.Label(frame_toolbar_table, text=f'Таблица', style='Toolbar.TLabel').pack(side=TOP, padx=2, pady=2)
        add_button = ttk.Button(frame_toolbar_table, image=self.plus_icon_pil, command=add_row_func)
        add_button.pack(side=LEFT, padx=3, pady=4)
        Hovertip(anchor_widget=add_button, text='Добавить строку в таблицу', hover_delay=400)
        edit_button = ttk.Button(frame_toolbar_table, image=self.edit_icon_pil, command=edit_row_func)
        edit_button.pack(side=LEFT, padx=3, pady=4)
        Hovertip(anchor_widget=edit_button, text='Редактировать строку', hover_delay=400)
        delete_button = ttk.Button(frame_toolbar_table, image=self.delete_icon_pil, command=delete_row_func)
        delete_button.pack(side=LEFT, padx=3, pady=4)
        Hovertip(anchor_widget=delete_button, text='Удалить строку', hover_delay=400)
        download_excel_button = ttk.Button(frame_toolbar_table, image=self.excel_icon_pil,
                                           command=lambda: export_excel_table(table=self.current_table))
        download_excel_button.pack(side=LEFT, padx=3, pady=4)
        Hovertip(anchor_widget=download_excel_button, text='Выгрузить таблицу в Excel файл', hover_delay=400)

        ttk.Label(self.frame_toolbar_attachments, text=f'Вложения', style='Toolbar.TLabel').pack(side=TOP, padx=2,
                                                                                                 pady=2)
        new_attachment_button = ttk.Button(self.frame_toolbar_attachments, image=self.new_attachment_icon_pil,
                                           command=new_attachment_func)
        new_attachment_button.pack(side=LEFT, padx=3, pady=4)
        Hovertip(anchor_widget=new_attachment_button, text='Прикрепить изображение / документ', hover_delay=400)
        self.open_attachments_button = ttk.Button(self.frame_toolbar_attachments, image=self.attachments_icon_pil,
                                                  command=looking_attachments_func)
        self.open_attachments_button.pack(side=LEFT, padx=3, pady=4)
        Hovertip(anchor_widget=self.open_attachments_button, text='Просмотреть вложения', hover_delay=400)

        ttk.Label(frame_toolbar_searching, text=f'Поиск', style='Toolbar.TLabel').pack(side=TOP, padx=2, pady=2)
        searching_button = ttk.Button(frame_toolbar_searching, image=self.loupe_icon_pil,
                                      command=lambda: self.render_typical_additional_window(
                                          called_class=Searcher,
                                          window_name='Spare Parts Searching'))
        searching_button.pack(side=LEFT, padx=3, pady=4)
        Hovertip(anchor_widget=searching_button, text='Выполнить поиск какой либо запчасти', hover_delay=400)

        ttk.Label(frame_toolbar_info, text=f'Справка', style='Toolbar.TLabel').pack(side=TOP, padx=2, pady=2)
        ttk.Button(frame_toolbar_info, image=self.info_icon_pil,
                   command=lambda: self.render_typical_additional_window(
                       called_class=lambda: ReferenceInfo(app_info=True),
                       window_name='App Info')).pack(side=LEFT, padx=3, pady=4)
        ttk.Button(frame_toolbar_info, image=self.help_icon_pil, command=lambda: self.render_typical_additional_window(
            called_class=ReferenceInfo,
            window_name='Reference Information')).pack(side=LEFT, padx=3,
                                                       pady=4)
        get_info_log(user=self.user_name, message='Toolbar widgets were rendered',
                     func_name=self.render_toolbar.__name__, func_path=abspath(__file__))

    def render_widgets_main_menu(self):
        """
        Функция рендера всех виджетов главного меню приложения
        """
        ttk.Label(self.frame_header, image=self.image_logo_pil).pack(side=LEFT, pady=2)
        (Label(self.frame_header, text="Mold Shop Management", width=100,
               font=("Times", "20", "bold"), background="deepskyblue", foreground="gainsboro").
         pack(side=TOP, padx=100, pady=20))

        ttk.Button(
            self.frame_main_widgets, style='Menu.TButton', text='Перечень пресс-форм',
            command=self.get_molds_data
        ).grid(padx=30, pady=15, column=5, row=2)

        ttk.Button(
            self.frame_main_widgets, style='Menu.TButton',
            text='Перемещение пресс-форм',
            command=self.open_mold_scanning_window
        ).grid(padx=30, pady=15, column=7, row=2)

        ttk.Button(
            self.frame_main_widgets, style='Menu.TButton',
            text='Поиск',
            command=lambda: self.render_typical_additional_window(
                called_class=Searcher,
                window_name='Spare Parts Searching')
        ).grid(padx=30, pady=15, column=9, row=2)

        ttk.Button(
            self.frame_main_widgets, style='Menu.TButton',
            text='Дефектация',
            command=lambda: self.render_typical_additional_window(
                called_class=MinPartsReport,
                window_name='Min Parts Searching'
            )
        ).grid(padx=30, pady=15, column=5, row=3)

        ttk.Button(
            self.frame_main_widgets, style='Menu.TButton',
            text='Журнал склада',
            command=lambda: self.open_warehouse_history_window(consumption=True)
        ).grid(padx=30, pady=15, column=7, row=3)

        ttk.Button(
            self.frame_main_widgets, style='Menu.TButton',
            text='Отслеживание закупок',
            command=self.open_purchased_parts_window
        ).grid(padx=30, pady=15, column=9, row=3)

        ttk.Button(
            self.frame_main_widgets, style='Menu.TButton',
            text='Справка', command=lambda: self.render_typical_additional_window(
                called_class=ReferenceInfo,
                window_name='Reference Information')
        ).grid(padx=30, pady=15, column=5, row=4)

        ttk.Label(self.frame_body, image=self.image_body_pil).pack(fill=BOTH, side=RIGHT, pady=10)
        get_info_log(user=self.user_name, message='Main menu widgets were rendered',
                     func_name=self.render_widgets_main_menu.__name__, func_path=abspath(__file__))

    def clean_frames(self):
        """
        Функция очистки окна приложения перед выводом новых виджетов
        """
        self.tree.unbind("<Double-ButtonPress-1>")
        self.tree.unbind("<Return>")
        self.tree.pack_forget()
        self.frame_header.pack_forget()
        self.frame_main_widgets.pack_forget()
        self.frame_body.pack_forget()
        self.frame_bottom.pack_forget()
        self.sort_status = None
        if self.frame_toolbar:
            self.frame_toolbar.pack_forget()
        if self.scroll_y:
            self.scroll_y.pack_forget()
        self.remove_listeners()

    def open_main_menu(self):
        """
        Вывод главного меню приложения
        """
        # Очистка контейнеров от старых виджетов
        self.clean_frames()
        # Обновление контейнеров
        self.frame_header = Frame(self, relief=RIDGE)
        self.frame_header.pack()
        self.frame_main_widgets = Frame(self)
        self.frame_main_widgets.pack(anchor=N)
        self.frame_body = Frame(self)
        self.frame_body.pack()
        # Рендер виджетов главного меню
        self.render_widgets_main_menu()

    def render_widgets_molds_list(self):
        """
        Функция рендера всех виджетов окна приложения в режиме просмотра перечня всех пресс-форм
        """
        molds_data_editing_access = user_data.get('molds_and_boms_data_changing')
        # Рендер панели инструментов
        self.render_toolbar(back_func=self.open_main_menu, add_row_func=lambda: self.render_typical_additional_window(
            called_class=EditedMold, window_name='New Mold Information', access=molds_data_editing_access,
            callback_function=self.get_molds_data),
                            edit_row_func=self.render_mold_edition_window,
                            delete_row_func=lambda: self.delete_selected_table_row('All_molds_data', 'MOLD_NUMBER'),
                            new_attachment_func=self.upload_attachment,
                            looking_attachments_func=lambda: self.render_typical_additional_window(
                                called_class=lambda: Attachment(
                                    mold_number=self.get_selected_row_data()[0]),
                                window_name='Attachments'))
        # Объявление основного и вложенных контейнеров для виджетов
        self.frame_main_widgets = Frame(self, relief=RIDGE)
        self.frame_main_widgets.pack(fill=X, expand=True)
        main_sub_frame = Frame(self.frame_main_widgets)
        main_sub_frame.pack(side=LEFT, pady=2, padx=2)
        picture_subframe = Frame(self.frame_main_widgets)
        picture_subframe.pack(side=RIGHT, pady=2, padx=2)
        title_frame = Frame(main_sub_frame)
        title_frame.pack(fill=BOTH, expand=True)
        description_frame = Frame(main_sub_frame)
        description_frame.pack(fill=BOTH, expand=True)
        bom_frame = ttk.LabelFrame(main_sub_frame, text='BOM', relief=RIDGE)
        bom_frame.pack(side=LEFT, padx=10, pady=1)
        sort_frame = ttk.LabelFrame(main_sub_frame, text='Сортировка п/ф', relief=RIDGE)
        sort_frame.pack(side=LEFT, padx=0, pady=1)
        # Рендер виджетов
        (ttk.Label(title_frame, text='Сводный перечень п/ф на балансе компании', style='Title.TLabel')
         .pack(side=LEFT, padx=8, pady=2))
        ttk.Label(description_frame, text='*********', style='Regular.TLabel').pack(side=LEFT, padx=8, pady=2)
        ttk.Label(bom_frame, text='Введите номер пресс-формы:', style='Regular.TLabel',
                  font=('Times', '9', 'normal')).pack(side=LEFT, padx=6, pady=5)

        self.mold_number_entry_field = ttk.Entry(bom_frame, font=('Times', '11', 'normal'))
        self.mold_number_entry_field.pack(side=LEFT, padx=6, pady=5)

        btn_bom = ttk.Button(
            bom_frame, text='Открыть', style='Regular.TButton',
            command=lambda: self.open_bom(mold_number=self.mold_number_entry_field.get())
        )
        btn_bom.pack(side=LEFT, padx=6, pady=5)

        btn_upload_bom = ttk.Button(
            bom_frame, text='Загрузить', style='Regular.TButton', width=10,
            command=lambda: render_window_for_selection_bom_type(window_title='New BOM Uploading',
                                                                 callback_func=self.upload_new_bom)
        )
        btn_upload_bom.pack(padx=6, pady=5, side=LEFT)
        Hovertip(anchor_widget=btn_upload_bom,
                 text='Для загрузки в систему нового BOM (спецификации) используйте шаблон таблицы, который можно '
                      'скачать.'
                      '\nЧтобы загрузка произошла успешно убедитесь, что пресс-форма под таким номером уже имеется в '
                      'общем'
                      '\nперечне и название файла полностью совпадает с номером пресс-формы. Например: название файла'
                      '\n"1981-A.xlsx" и номер пресс-формы "1981-A"',
                 hover_delay=400)

        btn_bom_template_download = ttk.Button(
            bom_frame, text='Скачать шаблон', style='Regular.TButton', width=14,
            command=lambda: render_window_for_selection_bom_type(window_title='BOM Template Downloading',
                                                                 callback_func=self.save_bom_parts_template)
        )
        btn_bom_template_download.pack(padx=6, pady=5, side=LEFT)
        Hovertip(anchor_widget=btn_bom_template_download,
                 text='Загрузка шаблонной таблицы для формирования BOM для последующей её загрузке в приложение',
                 hover_delay=400)

        btn_delete_bom = ttk.Button(
            bom_frame, text='Удалить', style='Regular.TButton',
            command=lambda: render_window_for_selection_bom_type(window_title='Удаление BOM',
                                                                 callback_func=self.delete_selected_bom)
        )
        btn_delete_bom.pack(padx=6, pady=5, side=LEFT)
        Hovertip(anchor_widget=btn_delete_bom,
                 text='Удаление BOM относящегося к выбранной пресс-форме',
                 hover_delay=400)

        ttk.Button(
            sort_frame, text='Все', style='Green.TButton',
            command=lambda: self.get_molds_data(sort_status=False)
        ).pack(padx=6, pady=5, side=LEFT)

        ttk.Button(
            sort_frame, text='Активные', style='Yellow.TButton',
            command=lambda: self.get_molds_data(sort_status='IN')
        ).pack(padx=6, pady=5, side=LEFT)

        ttk.Button(
            sort_frame, text='Не активные', style='Coral.TButton',
            command=lambda: self.get_molds_data(sort_status='OUT')
        ).pack(padx=6, pady=5, side=LEFT)

        ttk.Button(
            sort_frame, text='ТО2', style='Gold.TButton',
            command=lambda: self.get_molds_data(sort_status='IN SERVICE')
        ).pack(padx=6, pady=5, side=LEFT)

        ttk.Label(picture_subframe, image=self.image_mold_pil).pack(side=RIGHT, pady=1)

        get_info_log(user=self.user_name, message='Mold list widgets were rendered',
                     func_name=self.render_widgets_molds_list.__name__, func_path=abspath(__file__))

    def render_widgets_mold_scanning_mode(self):
        """
        Функция рендера всех виджетов окна приложения в режиме изменения статуса пресс-формы и просмотра
        журнала перемещений
        """
        # Рендер панели инструментов
        self.render_toolbar(back_func=self.open_main_menu)
        # Объявление основного и вложенных контейнеров для виджетов
        self.frame_main_widgets = Frame(self, relief=RIDGE)
        self.frame_main_widgets.pack(fill=X)
        main_sub_frame = Frame(self.frame_main_widgets)
        main_sub_frame.pack(side=LEFT, pady=2, padx=2)
        picture_subframe = Frame(self.frame_main_widgets)
        picture_subframe.pack(side=RIGHT, pady=2, padx=2)
        title_frame = Frame(main_sub_frame)
        title_frame.pack(fill=BOTH, expand=True)
        description_frame = Frame(main_sub_frame)
        description_frame.pack(fill=BOTH, expand=True)
        status_frame = ttk.LabelFrame(main_sub_frame, text='Статус пресс-формы', relief=RIDGE)
        status_frame.pack(side=LEFT, padx=4, pady=2)
        label_frame = Frame(main_sub_frame)
        label_frame.pack(side=LEFT, padx=25, pady=2)
        # Рендер виджетов
        (ttk.Label(title_frame, text='Перемещение пресс-форм', style='Title.TLabel')
         .pack(side=LEFT, padx=8, pady=2))
        ttk.Label(description_frame, text='*********', style='Regular.TLabel').pack(side=LEFT, padx=8, pady=2)
        ttk.Label(status_frame, text='Выберите необходимый статус куда будет перемещена п/ф').pack(side=TOP, padx=3,
                                                                                                   pady=2)

        ttk.Button(
            status_frame, text='ТО2', style='Regular.TButton',
            command=lambda: self.open_qr_window(next_status='IN SERVICE')
        ).pack(side=LEFT, padx=3, pady=3)

        ttk.Button(
            status_frame, text='Производство', style='Regular.TButton', width=12,
            command=lambda: self.open_qr_window(next_status='IN')
        ).pack(side=LEFT, padx=3, pady=3)

        ttk.Button(
            status_frame, text='Хранение', style='Regular.TButton',
            command=lambda: self.open_qr_window(next_status='OUT')
        ).pack(side=LEFT, padx=3, pady=3)

        ttk.Label(label_frame,
                  text='"IN" - П/Ф в производстве; "IN SERVICE" - П/Ф на обслуживании; "OUT" - П/Ф на хранении').pack(
            side=LEFT, padx=5)
        ttk.Label(picture_subframe, image=self.image_scanner_pil).pack(side=RIGHT, pady=1)

        get_info_log(user=self.user_name, message='Mold scanning mode widgets were rendered',
                     func_name=self.render_widgets_mold_scanning_mode.__name__, func_path=abspath(__file__))

    def render_widgets_warehouse_mode(self, consumption: bool = None):
        """
        Функция рендера всех виджетов окна приложения в режиме изменения статуса пресс-формы и просмотра журнала
        перемещений
        :param consumption: Булево значение, которое принимает значение True когда выбран расход
        (взятие запчастей со склада)
        """
        define_title: Callable = lambda: 'История расходов склада пресс-форм' if consumption \
            else 'История приходов склада пресс-форм'
        # Рендер панели инструментов
        self.render_toolbar(back_func=self.open_main_menu)
        # Объявление основного и вложенных контейнеров для виджетов
        self.frame_main_widgets = Frame(self, relief=RIDGE)
        self.frame_main_widgets.pack(fill=X)
        main_sub_frame = Frame(self.frame_main_widgets)
        main_sub_frame.pack(side=LEFT, pady=2, padx=2)
        picture_subframe = Frame(self.frame_main_widgets)
        picture_subframe.pack(side=RIGHT, pady=2, padx=2)
        title_frame = Frame(main_sub_frame)
        title_frame.pack(fill=BOTH, expand=True)
        description_frame = Frame(main_sub_frame)
        description_frame.pack(fill=BOTH, expand=True)
        status_frame = ttk.LabelFrame(main_sub_frame, text='Тип журнала склада', relief=RIDGE)
        status_frame.pack(side=LEFT, padx=4, pady=2)
        # Рендер виджетов
        (ttk.Label(title_frame, text=define_title(), style='Title.TLabel')
         .pack(side=LEFT, padx=8, pady=2))
        ttk.Label(description_frame, text='*********', style='Regular.TLabel').pack(side=LEFT, padx=8, pady=2)

        ttk.Button(
            status_frame, text='Приходы', style='Regular.TButton',
            command=self.open_warehouse_history_window
        ).pack(side=LEFT, padx=3, pady=3)

        ttk.Button(
            status_frame, text='Расходы', style='Regular.TButton',
            command=lambda: self.open_warehouse_history_window(consumption=True)
        ).pack(side=LEFT, padx=3, pady=3)

        ttk.Label(picture_subframe, image=self.image_rack_pil).pack(side=RIGHT, pady=1)

        get_info_log(user=self.user_name, message='Warehouse mode widgets were rendered',
                     func_name=self.render_widgets_warehouse_mode.__name__, func_path=abspath(__file__))

    def render_widgets_purchased_parts_mode(self):
        """
        Функция рендера всех виджетов окна приложения в режиме изменения статуса пресс-формы и просмотра журнала
        перемещений
        """
        # Рендер панели инструментов
        self.render_toolbar(back_func=self.open_main_menu)
        # Объявление основного и вложенных контейнеров для виджетов
        self.frame_main_widgets = Frame(self, relief=RIDGE)
        self.frame_main_widgets.pack(fill=X)
        main_sub_frame = Frame(self.frame_main_widgets)
        main_sub_frame.pack(side=LEFT, pady=2, padx=2)
        picture_subframe = Frame(self.frame_main_widgets)
        picture_subframe.pack(side=RIGHT, pady=2, padx=2)
        title_frame = Frame(main_sub_frame)
        title_frame.pack(fill=BOTH, expand=True)
        description_frame = Frame(main_sub_frame)
        description_frame.pack(fill=BOTH, expand=True)
        spareparts_frame = ttk.LabelFrame(main_sub_frame, text='Запчасти', relief=RIDGE)
        spareparts_frame.pack(side=LEFT, padx=4, pady=2)
        # Рендер виджетов
        (ttk.Label(title_frame, text='Отслеживание закупленных запчастей', style='Title.TLabel')
         .pack(side=LEFT, padx=8, pady=2))
        ttk.Label(description_frame, text='*********', style='Regular.TLabel').pack(side=LEFT, padx=8, pady=2)

        ttk.Button(
            spareparts_frame, text='Добавить закупаемые запчасти', style='Regular.TButton', width=30,
            command=self.update_window_with_new_parts,
        ).pack(side=LEFT, padx=3, pady=3)

        ttk.Button(
            spareparts_frame, text='Сделать выгрузку для таможни', style='Regular.TButton', width=30,
            command=lambda: self.render_typical_additional_window(called_class=CustomsReport,
                                                                  window_name='Customs report creation')
        ).pack(side=LEFT, padx=3, pady=3)

        ttk.Label(picture_subframe, image=self.image_rack_pil).pack(side=RIGHT, pady=1)

        get_info_log(user=self.user_name, message='Purchased parts mode widgets were rendered',
                     func_name=self.render_widgets_purchased_parts_mode.__name__, func_path=abspath(__file__))

    def render_widgets_selected_bom(self):
        """
        Функция рендера всех виджетов окна приложения в режиме просмотра BOM (спецификации) пресс-формы
        """
        define_table_name: Callable = lambda: f'BOM_HOT_RUNNER_{self.mold_number}' if self.hot_runner_bom \
            else f'BOM_{self.mold_number}'
        # Рендер панели инструментов
        self.render_toolbar(back_func=self.get_molds_data,
                            add_row_func=lambda: self.render_typical_additional_window(
                                called_class=lambda: EditedBOM(self.mold_number, define_table_name()),
                                window_name='New Spare Part Information',
                                access=user_data.get('molds_and_boms_data_changing'),
                                callback_function=lambda: self.open_bom(
                                    self.mold_number)),
                            edit_row_func=self.render_bom_edition_window,
                            delete_row_func=lambda: self.delete_selected_table_row(define_table_name(), 'NUMBER'),
                            new_attachment_func=lambda: self.upload_attachment(bom_part=True),
                            looking_attachments_func=self.render_attachments_window)
        # Объявление основного и вложенных контейнеров для виджетов
        self.frame_main_widgets = Frame(self, relief=RIDGE)
        self.frame_main_widgets.pack(fill=X)
        main_sub_frame = Frame(self.frame_main_widgets)
        main_sub_frame.pack(fill=X, side=LEFT, pady=2, padx=2)
        picture_subframe = Frame(self.frame_main_widgets)
        picture_subframe.pack(side=RIGHT, pady=2, padx=2)
        title_frame = Frame(main_sub_frame)
        title_frame.pack(fill=BOTH, expand=True)
        description_frame = Frame(main_sub_frame)
        description_frame.pack(fill=BOTH, expand=True)
        bom_frame = ttk.LabelFrame(main_sub_frame, text='BOM', relief=RIDGE)
        bom_frame.pack(side=LEFT, padx=10, pady=1)
        sort_frame = ttk.LabelFrame(main_sub_frame, text='Сортировка по количеству', relief=RIDGE)
        sort_frame.pack(side=LEFT, padx=0, pady=1)
        stock_frame = ttk.LabelFrame(main_sub_frame, text='Склад', relief=RIDGE)
        stock_frame.pack(side=LEFT, padx=5, pady=1)
        # Рендер виджетов
        molds_data = table_funcs.TableInDb('All_molds_data', 'Database')
        mold_info = molds_data.get_table(type_returned_data='dict', first_param='MOLD_NUMBER',
                                         first_value=self.mold_number, last_string=True)
        (ttk.Label(title_frame, text=f'BOM для пресс-формы № {self.mold_number} '
                                     f'({"Hot Runner" if self.hot_runner_bom else "Mold"})', style='Title.TLabel')
         .pack(side=LEFT, padx=8, pady=4))
        (ttk.Label(description_frame,
                   text=f'Проект: {mold_info.get("MOLD_NAME")} | Горячий канал: {mold_info.get("HOT_RUNNER_NUMBER")} | '
                        f'Год: {mold_info.get("RELEASE_YEAR")} | Кол-во гнёзд: {mold_info.get("CAVITIES_QUANTITY")}',
                   style='Regular.TLabel', font=('Arial', '9', 'normal'))
         .pack(side=LEFT, padx=8, pady=2))

        ttk.Button(
            bom_frame, text='Пресс-форма', style='Regular.TButton', width=12,
            command=lambda: self.open_bom(mold_number=self.mold_number)
        ).pack(side=LEFT, padx=6, pady=5)

        ttk.Button(
            bom_frame, text='Горячий канал', style='Regular.TButton', width=13,
            command=lambda: self.open_bom(mold_number=self.mold_number, hot_runner=True)
        ).pack(padx=6, pady=5, side=LEFT)

        ttk.Button(
            sort_frame, text='Все', style='Green.TButton',
            command=lambda: self.open_bom(mold_number=self.mold_number, hot_runner=self.hot_runner_bom)
        ).pack(padx=6, pady=5, side=LEFT)

        ttk.Button(
            sort_frame, text='В наличие', style='Yellow.TButton',
            command=lambda: self.open_bom(mold_number=self.mold_number, sort_status='В наличие',
                                          hot_runner=self.hot_runner_bom)
        ).pack(padx=6, pady=5, side=LEFT)

        ttk.Button(
            sort_frame, text='Отсутствующие', style='Coral.TButton',
            width=18,
            command=lambda: self.open_bom(mold_number=self.mold_number, sort_status='Отсутствующие',
                                          hot_runner=self.hot_runner_bom)
        ).pack(padx=6, pady=5, side=LEFT)

        ttk.Button(
            sort_frame, text='Меньше минимума', style='Gold.TButton',
            width=18,
            command=lambda: self.open_bom(mold_number=self.mold_number, sort_status='Меньше минимума',
                                          hot_runner=self.hot_runner_bom)
        ).pack(padx=6, pady=5, side=LEFT)

        ttk.Button(
            stock_frame, text='Списание', style='Regular.TButton', width=9,
            command=lambda: self.open_parts_quantity_changing_window(consumption=True)
        ).pack(padx=6, pady=5, side=LEFT)

        ttk.Button(
            stock_frame, text='Приход', style='Regular.TButton',
            command=self.open_parts_quantity_changing_window
        ).pack(padx=6, pady=5, side=LEFT)

        ttk.Button(
            stock_frame, text='Журнал', style='Regular.TButton',
            command=lambda: self.open_warehouse_history_window(consumption=True)
        ).pack(padx=6, pady=5, side=LEFT)

        ttk.Label(picture_subframe, image=self.image_spare_part_pil).pack(side=RIGHT, pady=1)

        get_info_log(user=self.user_name, message='BOM widgets were rendered',
                     func_name=self.render_widgets_selected_bom.__name__, func_path=abspath(__file__))

    def get_selected_row_data(self) -> list[Any] | Literal[""]:
        """
        Функция возврата значений выделенной строки пользователем в таблице
        :return: Список значений табличной строки
        """
        selected_rows = self.tree.selection()
        first_row_id = selected_rows[0]
        item = self.tree.item(first_row_id)
        return item['values']

    def on_clicked_or_pressed_table_row(self, event):
        """
        Обработчик события двойного нажатия мыши или клавиши Enter на выделенную строку таблицы
        :param event: Обрабатываемое событие
        """
        self.open_bom()
        self.event = event

    def on_selected_table_row(self, event):
        """
        Обработчик события нажатия мыши на выделенную строку таблицы
        :param event: Обрабатываемое событие
        """
        row_data = self.get_selected_row_data()
        if self.mold_number:
            part_number = row_data[0]
            attachment = (Attachment(
                mold_number=self.mold_number,
                part_number=part_number,
                hot_runner=self.hot_runner_bom,
                checking=True)
                          .search_attachments())
        else:
            mold_number = row_data[0]
            attachment = (Attachment(
                mold_number=mold_number,
                checking=True)
                          .search_attachments())

        if attachment:
            self.open_attachments_button.config(style='GreenToolbar.TButton')
        else:
            self.open_attachments_button.config(style='RedToolbar.TButton')
            self.event = event

    def on_clicked_or_pressed_purchased_parts_table_row(self, event):
        """
        Обработчик события двойного нажатия мыши или клавиши Enter на выделенную строку таблицы
        :param event: Обрабатываемое событие
        """
        part_info = self.get_selected_row_data()
        purchase_number = part_info[0]
        mold_number = part_info[2]
        part_number = part_info[4]
        part_name = part_info[5]
        status = part_info[8]
        comment = part_info[9]
        purchased_cnt = part_info[7]

        self.render_typical_additional_window(
            called_class=lambda: PurchasedPart(purchase_number=purchase_number,
                                               mold_number=mold_number,
                                               part_number=part_number,
                                               part_name=part_name,
                                               status=status,
                                               comment=comment,
                                               purchased_cnt=purchased_cnt),
            window_name='Purchased Part Changing',
            callback_function=self.open_purchased_parts_window)
        self.event = event

    def render_table(self, columns_sizes: dict):
        """
        Вывод таблицы в отображаемом окне приложения на основе полученных данных, которые были предварительно
        загружены из таблицы базы данных
        :param columns_sizes: Информация о размере каждого столбца таблицы
        """
        # Определение столбцов
        columns = self.current_table[0]
        # Определение таблицы
        self.tree = ttk.Treeview(columns=columns, show="headings")
        # Определение заголовков
        for col_name in columns:
            self.tree.heading(col_name, text=col_name)
        # Настройка столбцов
        for col_num, col_size in columns_sizes.items():
            self.tree.column(column=col_num, stretch=YES, width=col_size)
        # Добавление данных в отображаемую таблицу. Чтобы последняя добавленная информацию отображалась необходимо
        # прежде сделать реверсию массива данных
        reversed_data = list(reversed(self.current_table[1:]))
        for row in reversed_data:
            self.tree.insert("", END,
                             values=['-' if (cell is None or cell == '' or cell == ' ') else cell for cell in row])
        # Добавление вертикальной прокрутки
        self.scroll_y = tkinter.Scrollbar(orient=tkinter.VERTICAL, command=self.tree.yview)
        self.tree.configure(yscrollcommand=self.scroll_y.set)
        self.scroll_y.pack(side=RIGHT, fill='y')
        self.tree.pack(fill=BOTH, expand=1)

        ttk.Label(self.frame_bottom, text='Page 1', style='Regular.TLabel').pack(side=TOP, pady=5)

        get_info_log(user=user_data.get('user_name'), message='Table was rendered',
                     func_name=self.render_table.__name__, func_path=abspath(__file__))

    def get_molds_data(self, sort_status: str | bool = None):
        """
        Функция вывода перечня всех пресс-форм в табличном виде в окне приложения
        :param sort_status: Текстовое значение указывающее на необходимость проведения сортировки перечня п/ф
            по определённому критерию ("IN"; "OUT"; "IN SERVICE").
        """
        # Выгрузка из базы данных перечня пресс-форм, сортировка и запись в переменные класса
        try:
            self.sort_molds()
        except (sqlite3.OperationalError, AttributeError):
            messagebox.showerror(title='Ошибка', message='Ошибка обработки базы данных')
        else:
            # Очистка области в окне приложения перед выводом новой таблицы
            self.clean_frames()
            self.mold_number = None
            if sort_status:
                self.sort_status = sort_status
                self.current_table = self.sorted_molds_data_tuple.get(sort_status)
            else:
                self.sort_status = None
            # Рендер кнопок для данного окна
            self.render_widgets_molds_list()
            # Рендер таблицы
            self.render_table(columns_sizes=columns_sizes_molds_table)
            # Объявление обработчиков событий на клик мышью или клавиши Enter
            self.tree.bind('<Double-ButtonPress-1>', self.on_clicked_or_pressed_table_row)
            self.tree.bind('<<TreeviewSelect>>', self.on_selected_table_row)
            self.tree.bind('<Return>', self.on_clicked_or_pressed_table_row)

    def remove_listeners(self):
        """
        Функция удаления обработчиков событий
        """
        for key in range(1, 6):
            self.master.unbind(f'{key}')

    def sort_molds(self):
        """
        Функция сортировки общего перечня пресс-форм в зависимости от её статуса и записи информации в переменные
        класса. Вызывается при открытии окна с перечнем п/ф. Функция необходима для минимизации количества вызовов
        базы данных.
        """
        # Перезапись переменных для обновления данных, которые будут получены из БД
        self.sorted_molds_data_tuple = {status: [columns_molds_table] for status in mold_statuses_list}
        self.sorted_molds_data_dict = {status: [] for status in mold_statuses_list}
        # Выгрузка таблицы из БД
        molds_data = table_funcs.TableInDb('All_molds_data', 'Database')
        self.all_molds_table_dict = molds_data.get_table(type_returned_data='dict')
        self.molds_list_data = molds_data.get_table(type_returned_data='tuple')
        self.current_table = molds_data.get_table(type_returned_data='tuple')
        self.current_table.insert(0, columns_molds_table)  # Вставка наименований столбцов в открываемую таблицу
        # Сортировка и запись в зависимости от статуса
        for num, mold_info in enumerate(self.all_molds_table_dict):
            status = mold_info.get('STATUS')
            self.sorted_molds_data_dict.get(status).append(mold_info)
            self.sorted_molds_data_tuple.get(status).append(self.molds_list_data[num])

    def sort_bom_parts(self, hot_runner: bool = None):
        """
        Функция сортировки перечня BOM в зависимости от имеющегося количества запчастей на складе.
        Вызывается при открытии окна с BOM. Функция необходима для минимизации количества вызовов базы данных.
        :param hot_runner: Булево значение, которое характеризует какой тип BOM был выбран
        (Пресс-форма или горячий канал)
        """
        # Перезапись переменных для обновления данных, которые будут получены из БД
        self.sorted_bom_tuple = {status: [columns_bom_parts_table] for status in part_statuses_list}
        self.sorted_bom_dict = {status: [] for status in part_statuses_list}
        # Выгрузка таблицы из БД
        define_table_name: Callable = lambda: f'BOM_HOT_RUNNER_{self.mold_number}' if hot_runner \
            else f'BOM_{self.mold_number}'
        bom = table_funcs.TableInDb(define_table_name(), 'Database')
        self.bom_table_dict = list(reversed(bom.get_table(type_returned_data='dict')))
        self.bom_data = list(reversed(bom.get_table(type_returned_data='tuple')))
        self.current_table = list(reversed(bom.get_table(type_returned_data='tuple')))
        self.current_table.insert(0, columns_bom_parts_table)
        # Сортировка и запись в зависимости от статуса
        for num, part_info in enumerate(self.bom_table_dict):
            # Проверка числовых значений и сортировка данных по условиям
            try:
                parts_quantity = int(part_info.get('PARTS_QUANTITY'))
            except TypeError:
                self.sorted_bom_dict.get('Отсутствующие').append(part_info)
                self.sorted_bom_tuple.get('Отсутствующие').append(self.bom_data[num])
            else:
                if parts_quantity > 0:
                    self.sorted_bom_dict.get('В наличие').append(part_info)
                    self.sorted_bom_tuple.get('В наличие').append(self.bom_data[num])
                else:
                    self.sorted_bom_dict.get('Отсутствующие').append(part_info)
                    self.sorted_bom_tuple.get('Отсутствующие').append(self.bom_data[num])

            try:
                min_percent = int(part_info.get('MIN_PERCENT'))
                parts_quantity = int(part_info.get('PARTS_QUANTITY'))
                pcs_in_mold = int(part_info.get('PCS_IN_MOLDS'))
                current_percent = int(parts_quantity) / int(pcs_in_mold) * 100
            except TypeError:
                pass
            else:
                if parts_quantity > 0:
                    if current_percent < min_percent:
                        self.sorted_bom_dict.get('Меньше минимума').append(part_info)
                        self.sorted_bom_tuple.get('Меньше минимума').append(self.bom_data[num])

    def open_bom(self, mold_number: str = None, hot_runner: bool = None, sort_status: str | bool = None):
        """
        Функция для вывода спецификации (BOM) пресс-формы в табличном виде в окне приложения
        :param sort_status: Значение, которое характеризует был ли выбран параметр сортировки для отображения таблицы
        :param hot_runner: Булево значение, которое характеризует какой тип BOM был выбран
        (Пресс-форма или горячий канал)
        :param mold_number: Номер пресс-формы полученный из строки ввода
        """
        self.mold_number_entry_field.delete(0, END)
        self.hot_runner_bom = hot_runner
        define_table_name: Callable = lambda: f'BOM_HOT_RUNNER_{mold_number}' if hot_runner else f'BOM_{mold_number}'
        if not mold_number:
            mold_number = self.get_selected_row_data()[0]
        try:
            # Выгрузка информации из базы данных
            if sort_status:
                self.sort_status = sort_status
                self.sort_bom_parts(hot_runner)
                self.current_table = self.sorted_bom_tuple.get(sort_status)
            else:
                self.mold_number = mold_number
                bom = table_funcs.TableInDb(define_table_name(), 'Database')
                self.current_table = [columns_bom_parts_table]
                self.current_table.extend(list(reversed(bom.get_table(type_returned_data='tuple'))))
        except sqlite3.OperationalError:
            if hot_runner:
                messagebox.showerror('Уведомление об ошибке', f'Спецификации по номеру "{mold_number}" не имеется')
                self.hot_runner_bom = None
            else:
                # рекурсия для открытия BOM на горячий канал если не нашлась таблица с основным BOM
                try:
                    self.open_bom(mold_number=mold_number, hot_runner=True, sort_status=sort_status)
                except sqlite3.OperationalError:
                    messagebox.showerror('Уведомление об ошибке', f'Спецификации по номеру "{mold_number}" не имеется')
                    self.hot_runner_bom = None
        else:
            # Очистка области в окне приложения перед выводом новой таблицы
            self.clean_frames()
            # Обновление обработчиков событий
            self.remove_listeners()
            self.sort_status = sort_status
            self.render_widgets_selected_bom()
            self.render_table(columns_sizes=columns_sizes_bom_parts_table)
            self.tree.bind('<<TreeviewSelect>>', self.on_selected_table_row)

    def open_mold_scanning_window(self):
        """
        Функция вывода окна с виджетами для смены статуса пресс-формы и таблицей с историей этих изменений
        """
        # Очистка области в окне приложения перед выводом новой таблицы
        self.clean_frames()
        # Формирование табличных данных
        moving_history = table_funcs.TableInDb('Molds_moving_history', 'Database')
        moving_history_data = moving_history.get_table(type_returned_data='tuple')
        self.current_table = []
        self.current_table = [columns_molds_moving_history_table if i == 0 else moving_history_data[i - 1]
                              for i in range(0, len(moving_history_data) + 1)]
        # Рендер виджетов
        self.render_widgets_mold_scanning_mode()
        # Определение размера столбцов таблицы
        self.render_table(columns_sizes=columns_sizes_moving_history_table)
        # Объявление обработчиков событий
        # self.add_listeners(funk_two=self.open_main_menu)

    def open_warehouse_history_window(self, consumption: bool = None):
        """
        Функция вывода окна с виджетами для смены статусов пресс-форм и таблицей с историей их изменений
        :param consumption: Булево значение, которое принимает значение True когда происходит расход
        """
        define_table_name: Callable = lambda: 'OUT_warehouse_history' if consumption else 'IN_warehouse_history'
        # Очистка области в окне приложения перед выводом новой таблицы
        self.clean_frames()
        # Формирование табличных данных
        warehouse_history = table_funcs.TableInDb(define_table_name(), 'Database')
        warehouse_history_data = warehouse_history.get_table(type_returned_data='tuple')
        self.current_table = []
        self.current_table = [columns_warehouse_table if i == 0 else warehouse_history_data[i - 1]
                              for i in range(0, len(warehouse_history_data) + 1)]
        # Рендер виджетов
        self.render_widgets_warehouse_mode(consumption)
        # Определение размера столбцов таблицы
        self.render_table(columns_sizes=columns_sizes_warehouse_table)

    def open_purchased_parts_window(self):
        """
        Функция вывода окна с виджетами для смены статуса пресс-формы и таблицей с историей этих изменений
        """
        # Очистка области в окне приложения перед выводом новой таблицы
        self.clean_frames()
        # Формирование табличных данных
        purchased_parts = table_funcs.TableInDb('Purchased_parts', 'Database')
        purchased_parts_data = purchased_parts.get_table(type_returned_data='tuple')
        self.current_table = []
        self.current_table = [columns_purchased_parts if i == 0 else purchased_parts_data[i - 1]
                              for i in range(0, len(purchased_parts_data) + 1)]
        # Рендер виджетов
        self.render_widgets_purchased_parts_mode()
        # Определение размера столбцов таблицы
        self.render_table(columns_sizes=columns_sizes_purchased_parts_table)
        # Объявление обработчиков событий на двойной клик мышью или клавиши Enter
        self.tree.bind('<Double-ButtonPress-1>', self.on_clicked_or_pressed_purchased_parts_table_row)
        self.tree.bind('<Return>', self.on_clicked_or_pressed_purchased_parts_table_row)

    def update_window_with_new_parts(self):
        """
        Функция обновления окна с закупаемыми запчастями при успешной загрузке новых данных из Иксель файла
        """
        if upload_purchased_parts():
            self.open_purchased_parts_window()

    def open_qr_window(self, next_status: str):
        """
        Рендер окна для получения сообщения от сканера QR кода
        :param next_status: Новое значение параметра STATUS из таблицы перечня всех пресс-форм, которое заменит старое 
        """
        if user_data.get('mold_status_changing') == 'True':
            qr_window = QRWindow(next_status)
            qr_window.iconbitmap(os.path.join('pics', 'artpack.ico'))
            qr_window.title('QR code getting')
            qr_window.resizable(False, False)
            qr_window.render_widgets_before_getting_code()
            if qr_window.changed_data:
                get_info_log(user=self.user_name, message='Status changing',
                             func_name=self.open_qr_window.__name__, func_path=abspath(__file__))
                self.tree.pack_forget()
                self.open_mold_scanning_window()

    def render_typical_additional_window(self, called_class: Callable, window_name: str, access: str = None,
                                         callback_function: Callable = None):
        """
        Функция создания дополнительного окна по шаблону
        :param access: Переменная характеризующая наличие доступа у пользователя для выбранного действия
        :param callback_function: Вызываемая функция в случае изменения каких либо данных после взаимодействия
        пользователя в открытом окне
        :param window_name: Название открываемого окна
        :param called_class: Вызываемый класс для создания его экземпляра
        """
        if access != 'False':
            window = called_class()
            window.title(window_name)
            window.iconbitmap(os.path.join('pics', 'artpack.ico'))
            window.resizable(False, False)
            window.render_widgets()
            try:
                if window.changed_data:
                    get_info_log(user=self.user_name, message='Successful data changing',
                                 func_name=self.render_typical_additional_window.__name__, func_path=abspath(__file__))
                    self.tree.pack_forget()
                    if callback_function:
                        callback_function()
            except AttributeError:
                pass
        else:
            messagebox.showerror(error_messages.get('access_denied').get('message_name'),
                                 error_messages.get('access_denied').get('message_body'))

    def render_mold_edition_window(self):
        """
        Рендер окна для редактирования данных о пресс-форме
        """
        mold_data_edition_access = user_data.get('molds_and_boms_data_changing')
        mold_number = self.mold_number_entry_field.get()
        if not mold_number:
            mold_number = self.get_selected_row_data()[0]
        # Если получен номер элемента таблицы, тогда будет вызвано окно для взаимодействия с пользователем
        if mold_number:

            try:
                # Выгрузка информации о пресс-форме из базы данных
                all_molds_data = table_funcs.TableInDb('All_molds_data', 'Database')
                mold_data = all_molds_data.get_table(type_returned_data='dict', first_param='MOLD_NUMBER',
                                                     first_value=mold_number, last_string=True)
            except (sqlite3.OperationalError, IndexError):
                messagebox.showerror('Уведомление об ошибке', f'Данных о пресс-форме '
                                                              f'"{mold_number}" не имеется')
                get_warning_log(user=self.user_name, message='sqlite3.OperationalError',
                                func_name=self.render_mold_edition_window.__name__, func_path=abspath(__file__))
            else:
                self.render_typical_additional_window(called_class=lambda: EditedMold(mold_data=mold_data),
                                                      window_name='Mold Data Edition', access=mold_data_edition_access,
                                                      callback_function=self.get_molds_data)

    def render_bom_edition_window(self):
        """
        Рендер окна для редактирования данных о запчастях открытого в приложении BOM
        """
        bom_edition_access = user_data.get('molds_and_boms_data_changing')
        define_table_name: Callable = lambda: f'BOM_HOT_RUNNER_{self.mold_number}' if self.hot_runner_bom \
            else f'BOM_{self.mold_number}'
        table_name = define_table_name()
        part_info = self.get_selected_row_data()
        part_number = part_info[0]
        part_name = part_info[1]
        # Если получен номер элемента таблицы, тогда будет вызвано окно для взаимодействия с пользователем
        if part_number:

            try:
                # Выгрузка информации о пресс-форме из базы данных
                bom = table_funcs.TableInDb(table_name, 'Database')
                part_data = bom.get_table(type_returned_data='dict', first_param='NUMBER',
                                          first_value=part_number, second_param='PART_NAME', second_value=part_name,
                                          last_string=True)
            except (sqlite3.OperationalError, IndexError):
                messagebox.showerror('Уведомление об ошибке', f'Данных о запчасти не имеется')
                get_warning_log(user=self.user_name, message='sqlite3.OperationalError',
                                func_name=self.render_bom_edition_window.__name__, func_path=abspath(__file__))
            else:
                self.render_typical_additional_window(
                    called_class=lambda: EditedBOM(part_data=part_data, mold_number=self.mold_number,
                                                   table_name=table_name),
                    window_name='BOM Edition', access=bom_edition_access,
                    callback_function=lambda: self.open_bom(self.mold_number))

    def render_attachments_window(self):
        """
        Рендер окна для просмотра прикреплённых вложенных файлов к номеру пресс-формы, либо номеру запчасти
        """
        define_table_name: Callable = lambda: f'BOM_HOT_RUNNER_{self.mold_number}' if self.hot_runner_bom \
            else f'BOM_{self.mold_number}'
        table_name = define_table_name()
        part_number = self.get_selected_row_data()[0]
        # Если получен номер элемента таблицы, тогда будет вызвано окно для взаимодействия с пользователем
        if part_number:
            self.render_typical_additional_window(called_class=lambda: Attachment(mold_number=self.mold_number,
                                                                                  part_number=part_number,
                                                                                  hot_runner=self.hot_runner_bom),
                                                  window_name='Attachments')

    def open_parts_quantity_changing_window(self, consumption: bool = None):
        """
        Вывод окна для взаимодействия со складом. Взятие каких либо запчастей, либо её приём на склад. 
        :param consumption: Булево значение, которое принимает значение True когда происходит расход
        """
        define_stock_changing_access: Callable = lambda: user_data.get('stock_changing_out') if consumption \
            else user_data.get('stock_changing_in')
        define_table_name: Callable = lambda: f'BOM_HOT_RUNNER_{self.mold_number}' if self.hot_runner_bom \
            else f'BOM_{self.mold_number}'
        table_name = define_table_name()
        part_info = self.get_selected_row_data()
        part_number = part_info[0]
        part_name = part_info[1]
        # Если получен номер элемента таблицы, тогда будет вызвано окно для взаимодействия с пользователем
        if part_number and define_stock_changing_access():

            try:
                # Выгрузка информации о пресс-форме из базы данных
                bom = table_funcs.TableInDb(table_name, 'Database')
                part_data = bom.get_table(type_returned_data='dict', first_param='NUMBER',
                                          first_value=part_number, second_param='PART_NAME', second_value=part_name,
                                          last_string=True)
            except (sqlite3.OperationalError, IndexError):
                messagebox.showerror('Уведомление об ошибке', f'Данных о запчасти не имеется')
                get_warning_log(user=self.user_name, message='sqlite3.OperationalError',
                                func_name=self.render_bom_edition_window.__name__, func_path=abspath(__file__))
            else:
                self.render_typical_additional_window(
                    called_class=lambda: Stock(part_data=part_data, mold_number=self.mold_number,
                                               consumption=consumption, table_name=table_name),
                    window_name='BOM Edition', access=define_stock_changing_access(),
                    callback_function=lambda: self.open_bom(self.mold_number, self.hot_runner_bom))

    def delete_selected_table_row(self, table_name: str, column_name: str):
        """
        Функция удаления строки из таблицы базы данных на основании выделенной строки
        :param column_name: Имя столбца / параметра по которому будет искаться строка для удаления
        :param table_name: Имя таблицы из базы данных
        """
        if user_data.get('molds_and_boms_data_changing') == 'True':
            message = "Вы уверены, что хотите удалить данную строку"
            if messagebox.askyesno(title='Подтверждение', message=message, parent=self):

                try:
                    number = self.get_selected_row_data()[0]
                    db = table_funcs.TableInDb(table_name, 'Database')
                    db.delete_data(column_name=column_name, value=number)
                except sqlite3.OperationalError:
                    messagebox.showerror('Уведомление об ошибке', 'Ошибка удаления. Обратитесь к администратору.')
                    get_info_log(user=self.user_name, message='sqlite3.OperationalError',
                                 func_name=self.delete_selected_table_row.__name__, func_path=abspath(__file__))
                else:
                    messagebox.showinfo('Уведомление', 'Удаление успешно произведено!')
                    get_info_log(user=self.user_name, message='Row was deleted from table',
                                 func_name=self.delete_selected_table_row.__name__, func_path=abspath(__file__))

                    if column_name == 'MOLD_NUMBER':

                        try:
                            bom = table_funcs.TableInDb(f'BOM_{number}', 'Database')
                            bom.delete_db_table()
                        except sqlite3.OperationalError:
                            pass
                        self.tree.pack_forget()
                        self.get_molds_data()
                    else:
                        self.tree.pack_forget()
                        self.open_bom(mold_number=self.mold_number, hot_runner=self.hot_runner_bom)
        else:
            messagebox.showerror(error_messages.get('access_denied').get('message_name'),
                                 error_messages.get('access_denied').get('message_body'))

    def delete_selected_bom(self, previous_window, hot_runner_bom: bool = None):
        """
        Функция удаления BOM таблицы из базы данных на основании выделенной строки
        :param previous_window: Предыдущее окно с выбором BOM для удаления
        :param hot_runner_bom: Булево значение, которое характеризует какой тип BOM был выбран
        (Пресс-форма или горячий канал)
        """
        # Закрытие предыдущего окна с выбором BOM для удаления
        previous_window.quit()
        previous_window.destroy()

        if user_data.get('molds_and_boms_data_changing') == 'True':
            mold_number = self.mold_number_entry_field.get()
            if not mold_number:
                mold_number = self.get_selected_row_data()[0]

            if mold_number:
                define_table: Callable = lambda: f'BOM_HOT_RUNNER_{mold_number}' if hot_runner_bom \
                    else f'BOM_{mold_number}'
                message = "Вы уверены, что хотите удалить данный BOM"
                if messagebox.askyesno(title='Подтверждение', message=message, parent=self):

                    try:
                        bom = table_funcs.TableInDb(define_table(), 'Database')
                        bom.delete_db_table()
                    except sqlite3.OperationalError:
                        messagebox.showerror('Уведомление об ошибке',
                                             'Ошибка удаления. Вероятнее всего данный BOM отсутствовал.')
                        get_warning_log(user=self.user_name, message='sqlite3.OperationalError',
                                        func_name=self.delete_selected_bom.__name__, func_path=abspath(__file__))
                    else:
                        messagebox.showinfo('Уведомление', 'Удаление BOM успешно произведено!')
                        get_info_log(user=self.user_name, message='BOM was deleted from table',
                                     func_name=self.delete_selected_bom.__name__, func_path=abspath(__file__))
        else:
            messagebox.showerror(error_messages.get('access_denied').get('message_name'),
                                 error_messages.get('access_denied').get('message_body'))

    def confirm_delete(self):
        """
        Функция вывода диалогового окна для запроса подтверждения закрытия окна
        """
        message = "Вы уверены, что хотите удалить данную строку"
        if messagebox.askyesno(title='Подтверждение', message=message, parent=self):
            return True

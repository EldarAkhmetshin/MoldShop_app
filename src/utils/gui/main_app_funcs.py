#!/usr/bin/env python
# -*- coding: utf-8 -*- #
import shutil
import sqlite3
import tkinter
import os
import pandas
import PIL.ImageTk
from tkinter import *
from tkinter import ttk, filedialog, messagebox
from os.path import abspath
from ttkthemes import ThemedStyle
from pandas import DataFrame
from openpyxl import Workbook
from PIL import Image, ImageTk
from tkinter.ttk import Frame
from idlelib.tooltip import Hovertip
from typing import Callable

from src.global_values import user_data
from src.data import mold_statuses_dict, mold_statuses_list, part_statuses_list, columns_warehouse_table, \
    columns_sizes_warehouse_table
from src.molds import get_data_from_excel_file, check_mold_number
from src.data import info_messages, error_messages, columns_molds_moving_history_table, columns_sizes_moving_history_table
from src.utils.gui.bom_edition_funcs import EditedBOM
from src.utils.gui.mold_status_changing_funcs import QRWindow
from src.utils.gui.spare_parts_searching_funcs import Searcher
from src.utils.logger.logs import get_info_log, get_warning_log
from src.utils.sql_database import table_funcs, new_tables
from src.utils.gui.mold_data_edition_funcs import EditedMold
from src.utils.gui.stock_changing_funcs import Stock
from src.utils.gui.attached_files_review_funcs import Attachment


class App(Frame):
    """
    Класс представляет набор функций для создания графического интерфейса основного окна приложения с помощью
    библиотеки Tkinter. А именно отрисовка / вывод различных изображений, таблиц, кнопок, полей ввода и т.д. 
    Помимо рендера виджетов, класс включает в себя функции для обработки отображаемой информации. В частности это 
    взимодействие с таблицами базы данных. 
    """

    def __init__(self):
        """
        self.tree: Контейнер для вывода информации в табличном виде
        self.frame_header: Контейнер для отображения шапки приложения
        self.frame_main_widgets: Контейнер для отображения кнопок и полей ввода ниже шапки приложения
        self.frame_body: Контейнер для отображения основной информации
        self.tracked_variable: Переменная для отслеживания записываемых данных в поле ввода
        self.hot_runner_bom: Булево значение, которое становится True когда выбирается пользователем таблица горячего канала
        self.sorted_bom_tuple: Словарь, содержащий данные (BOM выбранной пресс-формы) в кортежах для вывода отсортированной информации по разным признакам
        self.all_molds_table_dict:
        self.mold_number: Текстовое значение параметра "MOLD_NUMBER" из таблицы базы данных. Принимает значение выбранной пресс-формы пользователем в данный момент
        self.mold_number_entry_field: Поле ввода для открытия BOM по номеру пресс-формы
        self.current_table: Массив состоящий из кортежей с данными для вывода таблицы в окне приложения
        self.event: Обрабатываемое событие
        self.sort_status: Значение характеризующее выбранный параметр сортировки для вывода табличных данных
        self.sorted_molds_data_tuple: Словарь, содержащий данные (общий перечень пресс-форм) в кортежах для вывода отсортированной информации по разным признакам
        self.sorted_molds_data_dict: Словарь, содержащий данные (общий перечень пресс-форм) в словарях для вывода отсортированной информации по разным признакам
        self.molds_list_data:
        """
        super().__init__()
        self.hot_runner_bom = None
        self.sorted_bom_tuple = None
        self.all_molds_table_dict = None
        self.mold_number = None
        self.current_table = None
        self.mold_number_entry_field = None
        self.event = None
        # Объявление переменных изображений
        self.image_logo = Image.open(os.path.abspath(os.path.join('pics', 'company_logo.png'))) \
            .resize((150, 70))
        self.image_logo_pil = ImageTk.PhotoImage(self.image_logo)
        self.image_body = Image.open(os.path.abspath(os.path.join('pics', 'main__picture.png'))) \
            .resize((800, 500))
        self.image_body_pil = ImageTk.PhotoImage(self.image_body)
        self.main_menu_icon = Image.open(os.path.abspath(os.path.join('pics', 'main_menu.png'))) \
            .resize((20, 20))
        self.main_menu_icon_pil = ImageTk.PhotoImage(self.main_menu_icon)
        self.image_mold = Image.open(os.path.abspath(os.path.join('pics', 'mold.png'))) \
            .resize((144, 120))
        self.image_mold_pil = ImageTk.PhotoImage(self.image_mold)
        self.image_spare_part = Image.open(os.path.abspath(os.path.join('pics', 'spare_parts.png'))) \
            .resize((144, 120))
        self.image_spare_part_pil = ImageTk.PhotoImage(self.image_spare_part)
        self.image_scanner = Image.open(os.path.abspath(os.path.join('pics', 'scanner.png'))) \
            .resize((144, 120))
        self.image_scanner_pil = ImageTk.PhotoImage(self.image_scanner)
        self.image_rack = Image.open(os.path.abspath(os.path.join('pics', 'rack.png'))) \
            .resize((144, 120))
        self.image_rack_pil = ImageTk.PhotoImage(self.image_rack)
        self.back_icon = Image.open(os.path.abspath(os.path.join('pics', 'back.png'))) \
            .resize((20, 20))
        self.back_icon_pil = ImageTk.PhotoImage(self.back_icon)
        self.plus_icon = Image.open(os.path.abspath(os.path.join('pics', 'plus.png'))) \
            .resize((20, 20))
        self.plus_icon_pil = ImageTk.PhotoImage(self.plus_icon)
        self.delete_icon = Image.open(os.path.abspath(os.path.join('pics', 'delete.png'))) \
            .resize((20, 20))
        self.delete_icon_pil = ImageTk.PhotoImage(self.delete_icon)
        self.edit_icon = Image.open(os.path.abspath(os.path.join('pics', 'edit.png'))) \
            .resize((20, 20))
        self.edit_icon_pil = ImageTk.PhotoImage(self.edit_icon)
        self.excel_icon = Image.open(os.path.abspath(os.path.join('pics', 'excel.png'))) \
            .resize((20, 20))
        self.excel_icon_pil = ImageTk.PhotoImage(self.excel_icon)
        self.info_icon = Image.open(os.path.abspath(os.path.join('pics', 'info.png'))) \
            .resize((20, 20))
        self.info_icon_pil = ImageTk.PhotoImage(self.info_icon)
        self.help_icon = Image.open(os.path.abspath(os.path.join('pics', 'help.png'))) \
            .resize((20, 20))
        self.help_icon_pil = ImageTk.PhotoImage(self.help_icon)
        self.new_attachment_icon = Image.open(os.path.abspath(os.path.join('pics', 'new_attachment.png'))) \
            .resize((20, 20))
        self.new_attachment_icon_pil = ImageTk.PhotoImage(self.new_attachment_icon)
        self.attachments_icon = Image.open(os.path.abspath(os.path.join('pics', 'attachments.png'))) \
            .resize((20, 20))
        self.attachments_icon_pil = ImageTk.PhotoImage(self.attachments_icon)
        self.loupe_icon = Image.open(os.path.abspath(os.path.join('pics', 'loupe.png'))) \
            .resize((20, 20))
        self.loupe_icon_pil = ImageTk.PhotoImage(self.loupe_icon)
        # Объявление переменных будущих контейнеров для хранения виджетов
        self.tree = ttk.Treeview()
        self.frame_header = None
        self.frame_toolbar = None
        self.frame_main_widgets = None
        self.frame_body = None

        self.sort_status = None
        self.sorted_molds_data_tuple = None
        self.sorted_molds_data_dict = None
        self.molds_list_data = None

        self.key_two_func = None
        self.key_three_func = None
        self.key_four_func = None
        self.key_five_func = None

        self.init_gui()

    def init_gui(self):
        """
        Инициация окна приложения, а также основных контейнеров
        """
        self.master.title('MoldShop Management')
        self.pack(fill=BOTH)

        self.frame_header = Frame(self, relief=RAISED)
        self.frame_header.pack(fill=BOTH, expand=True)
        self.frame_main_widgets = Frame(self)
        self.frame_main_widgets.pack(anchor=N)
        self.frame_body = Frame(self)
        self.frame_body.pack(fill=BOTH, expand=True)

    def open_file_and_download(self, window: tkinter, hot_runner: bool = None):
        """
        Функция загрузки спецификации пресс-формы (BOM) из Иксель файла формата xlsx в таблицу базы данных
        :param window: дополнительное окно приложения
        :param hot_runner: Булево значение, которое характеризует какой тип BOM был выбран (Пресс-форма или горячий канал)
        """
        # Открытие диалогового окна для выбора файла пользователем с локальной директории компьютера,
        # c дальнейшим извлечением пути к выбранному файлу в виде строки
        try:
            window.quit()
            window.destroy()
            file_path = filedialog.askopenfile(
                filetypes=(('XLSX files', '*.xlsx'),)
            ).name
        except AttributeError:
            pass
        else:
            # Получение информации из Иксель файла типа xlsx
            if hot_runner:
                column_names, rows_data = get_data_from_excel_file(file_path=file_path,
                                                                   work_sheet_name='HOT_RUNNER')
            else:
                column_names, rows_data = get_data_from_excel_file(file_path=file_path, work_sheet_name='BOM')
            file_path = file_path.split('/')
            mold_number = file_path[-1].replace('.xlsx', '')
            # Поиск соответствия по номеру пресс-формы в общем перечне
            if check_mold_number(mold_number):
                # Сохранение информации в базе данных
                new_tables.create_bom_for_new_mold(mold_number=mold_number, rows_data=rows_data, hot_runner=hot_runner)
                # Рендер окна приложения с новой загруженной информацией
                info_message = info_messages.get('downloaded_bom').get('message_body')
                self.mold_number = mold_number
                self.open_bom(mold_number=mold_number)
                messagebox.showinfo(title=info_messages.get('downloaded_bom').get('message_name'),
                                    message=info_message.format(mold_number=mold_number))
            else:
                error_message = error_messages.get('not_downloaded_bom').get('message_body')
                messagebox.showerror(title=error_messages.get('not_downloaded_bom').get('message_name'),
                                     message=error_message.format(mold_number=mold_number))

    def upload_attachment(self, bom_part: bool = None, hot_runner: bool = None):
        """
        Функция загрузки вложения (какого либо файла) в директорию приложения для дальнейшего использования
        :param part_number: ID номер элемента BOM относящегося к определённой пресс-форме
        :param hot_runner: Булево значение, которое характеризует какой тип BOM был выбран (Пресс-форма или горячий канал)
        """
        # Открытие диалогового окна для выбора файла пользователем с локальной директории компьютера,
        # c дальнейшим извлечением пути к выбранному файлу в виде строки
        if bom_part:
            define_table_name: Callable = lambda: f'BOM_HOT_RUNNER_{self.mold_number}' if self.hot_runner_bom else f'BOM_{self.mold_number}'
            part_number = self.get_value_by_selected_row(define_table_name(), 'NUMBER')

        if not self.mold_number:
            self.mold_number = self.get_value_by_selected_row('All_molds_data', 'MOLD_NUMBER')
        if self.mold_number:
            try:
                file_path = filedialog.askopenfile().name
            except AttributeError:
                pass
            else:
                # Получение имени файла
                file_path_list = file_path.split('/')
                file_name = file_path_list[-1]
                # Создание нужной директории где будет размещён файл если она ещё не была создана
                define_folder: Callable = lambda: os.path.join('savings', 'attachments', self.mold_number) \
                    if not part_number \
                    else (
                    os.path.join('savings', 'attachments', self.mold_number, 'hot_runner_parts', part_number)
                    if hot_runner else os.path.join('savings', 'attachments', self.mold_number, 'mold_parts',
                                                    part_number))
                try:
                    os.mkdir(os.path.join('savings', 'attachments', self.mold_number))
                except FileExistsError:
                    pass
                try:
                    os.mkdir(os.path.join('savings', 'attachments', self.mold_number, 'hot_runner_parts'))
                    os.mkdir(os.path.join('savings', 'attachments', self.mold_number, 'mold_parts'))
                except FileExistsError:
                    pass
                if bom_part:
                    try:
                        os.mkdir(define_folder())
                    except FileExistsError:
                        pass

                # Копирование и вставка файла в директорию приложения
                define_path: Callable = lambda: os.path.join('savings', 'attachments', self.mold_number, file_name) \
                    if not bom_part \
                    else (
                    os.path.join('savings', 'attachments', self.mold_number, 'hot_runner_parts', part_number, file_name)
                    if hot_runner else os.path.join('savings', 'attachments', self.mold_number, 'mold_parts',
                                                    part_number, file_name))
                try:
                    shutil.copy2(file_path, os.path.abspath(define_path()))
                except IOError:
                    messagebox.showerror(title='Ошибка',
                                         message='Файл не удалось прикрепить. Обратитесь к администратору.')
                else:
                    messagebox.showinfo(title='Уведомление', message='Файл успешно прикреплён')
                    self.render_typical_additional_window(called_class=lambda: Attachment(
                        mold_number=self.mold_number, part_number=part_number, hot_runner=hot_runner),
                                                          window_name='Attachments')
        else:
            messagebox.showerror(title='Ошибка',
                                 message='Чтобы прикрепить файл выберите элемент из таблицы')

    def save_excel_table(self):
        """
        Функция сохранения открытой таблицы из окна приложения в Иксель файл формата xlsx
        """
        # Открытие диалогового окна для сохранения файла пользователем в локальной директории компьютера,
        # c дальнейшим извлечением пути к выбранному файлу в виде строки
        import openpyxl
        from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
        from openpyxl.utils import get_column_letter

        file_path = filedialog.asksaveasfilename(
            filetypes=(('XLSX files', '*.xlsx'),)
        )
        if file_path:
            file_path = file_path.replace('.xlsx', '')
            # Формирование Иксель файла типа xlsx
            table_length = len(self.current_table)
            saved_table = {}
            for num, column_name in enumerate(self.current_table[0]):
                saved_table[column_name] = tuple(self.current_table[i][num] for i in range(1, table_length))
            try:
                df = DataFrame(saved_table)
                df.to_excel(excel_writer=f'{file_path}.xlsx', sheet_name='Table', index=False)
                wb = openpyxl.load_workbook(f'{file_path}.xlsx')
                ws = wb['Table']
                dim_holder = DimensionHolder(worksheet=ws)

                for col in range(ws.min_column, ws.max_column + 1):
                    dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=20)

                ws.column_dimensions = dim_holder
                wb.save(f'{file_path}.xlsx')
            except Exception:
                messagebox.showerror(title='Уведомление об ошибке',
                                     message='Ошибка в записи файла.\nПовторите ещё раз, либо братитесь к администратору')
            else:
                messagebox.showinfo(title='Уведомление', message='Таблица успешно сохранена на Ваш компьютер')

    def render_toolbar(self, back_func: Callable, add_row_func: Callable = None, edit_row_func: Callable = None,
                       delete_row_func: Callable = None, new_attachment_func: Callable = None,
                       looking_attachments_func: Callable = None):
        """
        Рендер виджетов панели иснтрументов
        :param looking_attachments_func: Функция просмотра вложенных файлов
        :param new_attachment_func: Функция прикрепления файла
        :param back_func: Вызываемая функция при нажатии кнопки "Назад" 
        :param add_row_func: Функция добавления новой строки в открытой таблице
        :param edit_row_func: Функция редактирования выбранной строки в открытой таблице
        :param delete_row_func: Функция удаления выбранной строки в открытой таблице
        """
        # Объявление контейнеров
        self.frame_toolbar = ttk.Frame(self)
        self.frame_toolbar.pack(fill=BOTH, expand=True)
        frame_toolbar_menu = Frame(self.frame_toolbar, relief=RIDGE)
        frame_toolbar_menu.pack(side=LEFT, padx=0)
        frame_toolbar_table = Frame(self.frame_toolbar, relief=RIDGE)
        frame_toolbar_table.pack(side=LEFT, padx=0)
        frame_toolbar_attachments = Frame(self.frame_toolbar, relief=RIDGE)
        frame_toolbar_attachments.pack(side=LEFT, padx=0)
        frame_toolbar_searching = Frame(self.frame_toolbar, relief=RIDGE)
        frame_toolbar_searching.pack(side=LEFT, padx=0)
        frame_toolbar_picture = Frame(self.frame_toolbar, relief=RIDGE)
        frame_toolbar_picture.pack(side=LEFT, padx=0)
        frame_toolbar_info = Frame(self.frame_toolbar, relief=RIDGE)
        frame_toolbar_info.pack(side=RIGHT, padx=0)
        # Объявление виджетов и их рендер
        ttk.Label(frame_toolbar_menu, text=f'Меню', anchor=NW, style='Toolbar.TLabel').pack(expand=True, padx=2, pady=2)
        menu_button = ttk.Button(frame_toolbar_menu, image=self.main_menu_icon_pil,
                                 command=self.open_main_menu)
        menu_button.pack(side=LEFT, padx=3, pady=4)
        Hovertip(anchor_widget=menu_button, text='Главное меню', hover_delay=400)
        back_button = ttk.Button(frame_toolbar_menu, image=self.back_icon_pil,
                                 command=back_func)
        back_button.pack(side=LEFT, padx=3, pady=4)
        Hovertip(anchor_widget=back_button, text='Назад', hover_delay=400)

        ttk.Label(frame_toolbar_table, text=f'Таблица', style='Toolbar.TLabel').pack(side=TOP, padx=2, pady=2)
        add_button = ttk.Button(frame_toolbar_table, image=self.plus_icon_pil, command=add_row_func)
        add_button.pack(side=LEFT, padx=3, pady=4)
        Hovertip(anchor_widget=add_button, text='Добавить строку в таблицу', hover_delay=400)
        edit_button = ttk.Button(frame_toolbar_table, image=self.edit_icon_pil, command=edit_row_func)
        edit_button.pack(side=LEFT, padx=3, pady=4)
        Hovertip(anchor_widget=edit_button, text='Редактировать строку', hover_delay=400)
        delete_button = ttk.Button(frame_toolbar_table, image=self.delete_icon_pil, command=delete_row_func)
        delete_button.pack(side=LEFT, padx=3, pady=4)
        Hovertip(anchor_widget=delete_button, text='Удалить строку', hover_delay=400)
        download_excel_button = ttk.Button(frame_toolbar_table, image=self.excel_icon_pil,
                                           command=self.save_excel_table)
        download_excel_button.pack(side=LEFT, padx=3, pady=4)
        Hovertip(anchor_widget=download_excel_button, text='Выгрузить таблицу в Excel файл', hover_delay=400)

        ttk.Label(frame_toolbar_attachments, text=f'Вложения', style='Toolbar.TLabel').pack(side=TOP, padx=2, pady=2)
        new_attachment_button = ttk.Button(frame_toolbar_attachments, image=self.new_attachment_icon_pil,
                                           command=new_attachment_func)
        new_attachment_button.pack(side=LEFT, padx=3, pady=4)
        Hovertip(anchor_widget=new_attachment_button, text='Прикрепить изображение / документ', hover_delay=400)
        open_attacments_button = ttk.Button(frame_toolbar_attachments, image=self.attachments_icon_pil,
                                            command=looking_attachments_func)
        open_attacments_button.pack(side=LEFT, padx=3, pady=4)
        Hovertip(anchor_widget=open_attacments_button, text='Просмотреть вложения', hover_delay=400)

        ttk.Label(frame_toolbar_searching, text=f'Поиск', style='Toolbar.TLabel').pack(side=TOP, padx=2, pady=2)
        searching_button = ttk.Button(frame_toolbar_searching, image=self.loupe_icon_pil,
                                      command=lambda: self.render_typical_additional_window(
                                          called_class=Searcher,
                                          window_name='Spare Parts Searching'))
        searching_button.pack(side=LEFT, padx=3, pady=4)
        Hovertip(anchor_widget=searching_button, text='Выполнить поиск какой либо запчасти', hover_delay=400)

        ttk.Label(frame_toolbar_info, text=f'Справка', style='Toolbar.TLabel').pack(side=TOP, padx=2, pady=2)
        ttk.Button(frame_toolbar_info, image=self.info_icon_pil,
                   command=self.open_main_menu).pack(side=LEFT, padx=3, pady=4)
        ttk.Button(frame_toolbar_info, image=self.help_icon_pil, command=self.open_main_menu).pack(side=LEFT, padx=3,
                                                                                                   pady=4)

    def render_widgets_main_menu(self):
        """
        Функция рендера всех виджетов главного меню приложения
        """
        ttk.Label(self.frame_header, image=self.image_logo_pil).pack(side=LEFT, pady=2)
        (Label(self.frame_header, text="Mold Shop Management", width=100,
               font=("Times", "20", "bold"), background="deepskyblue", foreground="gainsboro").
         pack(side=TOP, padx=100, pady=20))

        ttk.Button(
            self.frame_main_widgets, style='Menu.TButton', text='Перечень пресс-форм',
            command=self.get_molds_data
        ).grid(padx=30, pady=15, column=5, row=2)

        ttk.Button(
            self.frame_main_widgets, style='Menu.TButton',
            text='Перемещение пресс-форм',
            command=self.open_mold_scanning_window
        ).grid(padx=30, pady=15, column=7, row=2)

        ttk.Button(
            self.frame_main_widgets, style='Menu.TButton',
            text='Поиск',
            command=lambda: self.render_typical_additional_window(
                called_class=Searcher,
                window_name='Spare Parts Searching')
        ).grid(padx=30, pady=15, column=9, row=2)

        ttk.Button(
            self.frame_main_widgets, style='Menu.TButton',
            text='Дефектация'
        ).grid(padx=30, pady=15, column=5, row=3)

        ttk.Button(
            self.frame_main_widgets, style='Menu.TButton',
            text='Журнал склада',
            command=lambda: self.open_warehouse_history_window(consumption=True)
        ).grid(padx=30, pady=15, column=7, row=3)

        ttk.Button(
            self.frame_main_widgets, style='Menu.TButton',
            text='Справка'
        ).grid(padx=30, pady=15, column=9, row=3)

        ttk.Label(self.frame_body, image=self.image_body_pil).pack(side=RIGHT, pady=27)

    def clean_frames(self):
        """
        Функция очистки окна приложения перед выводом новых виджетов
        """
        self.tree.unbind("<Double-ButtonPress-1>")
        self.tree.unbind("<Return>")
        self.tree.pack_forget()
        self.frame_header.pack_forget()
        self.frame_main_widgets.pack_forget()
        self.frame_body.pack_forget()
        self.sort_status = None
        if self.frame_toolbar:
            self.frame_toolbar.pack_forget()
        self.remove_listeners()

    def open_main_menu(self):
        """
        Вывод главного меню приложения
        """
        # Очистка контейнеров от старых виджетов
        self.clean_frames()
        # Обновление контейнеров
        self.frame_header = Frame(self, relief=RIDGE)
        self.frame_header.pack()
        self.frame_main_widgets = Frame(self)
        self.frame_main_widgets.pack(anchor=N)
        self.frame_body = Frame(self)
        self.frame_body.pack(fill=BOTH, expand=True)
        # Рендеров виджетов главного меню
        self.render_widgets_main_menu()

    def render_widgets_molds_list(self):
        """
        Функция рендера всех виджетов окна приложения в режиме просмотра перечня всех пресс-форм
        """
        molds_data_editing_access = user_data.get('molds_and_boms_data_changing')
        # Рендер панели инструментов
        self.render_toolbar(back_func=self.open_main_menu, add_row_func=lambda: self.render_typical_additional_window(
            called_class=EditedMold, window_name='New Mold Information', access=molds_data_editing_access, called_function=self.get_molds_data),
                            edit_row_func=self.render_mold_edition_window,
                            delete_row_func=lambda: self.delete_selected_table_row('All_molds_data', 'MOLD_NUMBER'),
                            new_attachment_func=self.upload_attachment,
                            looking_attachments_func=lambda: self.render_typical_additional_window(
                                called_class=lambda: Attachment(
                                    mold_number=self.get_value_by_selected_row('All_molds_data', 'MOLD_NUMBER')),
                                window_name='Attachments'))
        # Объявление основного и вложенных контейнеров для виджетов
        self.frame_main_widgets = Frame(self, relief=RIDGE)
        self.frame_main_widgets.pack(fill=X, expand=True)
        main_sub_frame = Frame(self.frame_main_widgets)
        main_sub_frame.pack(side=LEFT, pady=2, padx=2)
        picture_subframe = Frame(self.frame_main_widgets)
        picture_subframe.pack(side=RIGHT, pady=2, padx=2)
        title_frame = Frame(main_sub_frame)
        title_frame.pack(fill=BOTH, expand=True)
        description_frame = Frame(main_sub_frame)
        description_frame.pack(fill=BOTH, expand=True)
        bom_frame = ttk.LabelFrame(main_sub_frame, text='BOM', relief=RIDGE)
        bom_frame.pack(side=LEFT, padx=10, pady=1)
        sort_frame = ttk.LabelFrame(main_sub_frame, text='Сортировка п/ф', relief=RIDGE)
        sort_frame.pack(side=LEFT, padx=0, pady=1)
        # Рендер виджетов
        (ttk.Label(title_frame, text='Сводный перечень п/ф на балансе компании', style='Title.TLabel')
         .pack(side=LEFT, padx=8, pady=2))
        ttk.Label(description_frame, text='*********', style='Regular.TLabel').pack(side=LEFT, padx=8, pady=2)
        ttk.Label(bom_frame, text='Введите номер пресс-формы:', style='Regular.TLabel',
                  font=('Times', '9', 'normal')).pack(side=LEFT, padx=6, pady=5)

        self.mold_number_entry_field = ttk.Entry(bom_frame, font=('Times', '11', 'normal'))
        self.mold_number_entry_field.pack(side=LEFT, padx=6, pady=5)

        btn_bom = ttk.Button(
            bom_frame, text='Открыть', style='Regular.TButton',
            command=lambda: self.open_bom(mold_number=self.mold_number_entry_field.get())
        )
        btn_bom.pack(side=LEFT, padx=6, pady=5)

        btn_download_bom = ttk.Button(
            bom_frame, text='Загрузить', style='Regular.TButton',
            command=self.render_upload_bom_window
        )
        btn_download_bom.pack(padx=6, pady=5, side=LEFT)
        Hovertip(anchor_widget=btn_download_bom,
                 text='Для загрузки в систему нового BOM (спецификации) используйте шаблон таблицы "mold_bom_template.xlsx" '
                      '\nиз папки "templates". '
                      '\nЧтобы загрузка произошла успешно убедитесь, что пресс-форма под таким номером уже имеется в общем '
                      '\nперечне и название файла полностью совпадает с номером пресс-формы. Например: название файла'
                      '\n"1981-A.xlsx" и номер пресс-формы "1981-A"',
                 hover_delay=400)

        btn_delete_bom = ttk.Button(
            bom_frame, text='Удалить', style='Regular.TButton',
            command=self.delete_selected_bom
        )
        btn_delete_bom.pack(padx=6, pady=5, side=LEFT)
        Hovertip(anchor_widget=btn_delete_bom,
                 text='Удаление BOM относящегося к выбранной пресс-форме',
                 hover_delay=400)

        ttk.Button(
            sort_frame, text='Все', style='Green.TButton',
            command=lambda: self.get_molds_data(sort_status=False)
        ).pack(padx=6, pady=5, side=LEFT)

        ttk.Button(
            sort_frame, text='Активные', style='Yellow.TButton',
            command=lambda: self.get_molds_data(sort_status='IN')
        ).pack(padx=6, pady=5, side=LEFT)

        ttk.Button(
            sort_frame, text='Не активные', style='Coral.TButton',
            command=lambda: self.get_molds_data(sort_status='OUT')
        ).pack(padx=6, pady=5, side=LEFT)

        ttk.Button(
            sort_frame, text='ТО2', style='Gold.TButton',
            command=lambda: self.get_molds_data(sort_status='IN SERVICE')
        ).pack(padx=6, pady=5, side=LEFT)

        ttk.Label(picture_subframe, image=self.image_mold_pil).pack(side=RIGHT, pady=1)

        get_info_log(user=user_data.get('user_name'), message='Mold list widgets were rendered',
                     func_name=self.render_widgets_molds_list.__name__, func_path=abspath(__file__))

    def render_widgets_mold_scanning_mode(self):
        """
        Функция рендера всех виджетов окна приложения в режиме изменения статуса пресс-формы и просмотра журнала перемещений
        """
        # Рендер панели инструментов
        self.render_toolbar(back_func=self.open_main_menu)
        # Объявление основного и вложенных контейнеров для виджетов
        self.frame_main_widgets = Frame(self, relief=RIDGE)
        self.frame_main_widgets.pack(fill=X)
        main_sub_frame = Frame(self.frame_main_widgets)
        main_sub_frame.pack(side=LEFT, pady=2, padx=2)
        picture_subframe = Frame(self.frame_main_widgets)
        picture_subframe.pack(side=RIGHT, pady=2, padx=2)
        title_frame = Frame(main_sub_frame)
        title_frame.pack(fill=BOTH, expand=True)
        description_frame = Frame(main_sub_frame)
        description_frame.pack(fill=BOTH, expand=True)
        status_frame = ttk.LabelFrame(main_sub_frame, text='Статус пресс-формы', relief=RIDGE)
        status_frame.pack(side=LEFT, padx=4, pady=2)
        # Рендер виджетов
        (ttk.Label(title_frame, text='Перемещение пресс-форм', style='Title.TLabel')
         .pack(side=LEFT, padx=8, pady=2))
        ttk.Label(description_frame, text='*********', style='Regular.TLabel').pack(side=LEFT, padx=8, pady=2)
        ttk.Label(status_frame, text='Выберите необходимый статус куда будет перещена п/ф').pack(side=TOP, padx=3,
                                                                                                 pady=2)

        ttk.Button(
            status_frame, text='ТО2', style='Regular.TButton',
            command=lambda: self.open_qr_window(next_status='IN SERVICE')
        ).pack(side=LEFT, padx=3, pady=3)

        ttk.Button(
            status_frame, text='Производство', style='Regular.TButton',
            command=lambda: self.open_qr_window(next_status='IN')
        ).pack(side=LEFT, padx=3, pady=3)

        ttk.Button(
            status_frame, text='Хранение', style='Regular.TButton',
            command=lambda: self.open_qr_window(next_status='OUT')
        ).pack(side=LEFT, padx=3, pady=3)

        ttk.Label(picture_subframe, image=self.image_scanner_pil).pack(side=RIGHT, pady=1)

        get_info_log(user=user_data.get('user_name'), message='Mold scaning mode widgets were rendered',
                     func_name=self.render_widgets_molds_list.__name__, func_path=abspath(__file__))

    def render_widgets_warehouse_mode(self, consumption: bool = None):
        """
        Функция рендера всех виджетов окна приложения в режиме изменения статуса пресс-формы и просмотра журнала перемещений
        :param consumption: Булево значение, которое принимает значение True когда выбран расход (взятие запчастей со склада)
        """
        define_title: Callable = lambda: 'История расходов склада пресс-форм' if consumption \
            else 'История приходов склада пресс-форм'
        # Рендер панели инструментов
        self.render_toolbar(back_func=self.open_main_menu)
        # Объявление основного и вложенных контейнеров для виджетов
        self.frame_main_widgets = Frame(self, relief=RIDGE)
        self.frame_main_widgets.pack(fill=X)
        main_sub_frame = Frame(self.frame_main_widgets)
        main_sub_frame.pack(side=LEFT, pady=2, padx=2)
        picture_subframe = Frame(self.frame_main_widgets)
        picture_subframe.pack(side=RIGHT, pady=2, padx=2)
        title_frame = Frame(main_sub_frame)
        title_frame.pack(fill=BOTH, expand=True)
        description_frame = Frame(main_sub_frame)
        description_frame.pack(fill=BOTH, expand=True)
        status_frame = ttk.LabelFrame(main_sub_frame, text='Тип журнала склада', relief=RIDGE)
        status_frame.pack(side=LEFT, padx=4, pady=2)
        # Рендер виджетов
        (ttk.Label(title_frame, text=define_title(), style='Title.TLabel')
         .pack(side=LEFT, padx=8, pady=2))
        ttk.Label(description_frame, text='*********', style='Regular.TLabel').pack(side=LEFT, padx=8, pady=2)

        ttk.Button(
            status_frame, text='Приходы', style='Regular.TButton',
            command=self.open_warehouse_history_window
        ).pack(side=LEFT, padx=3, pady=3)

        ttk.Button(
            status_frame, text='Расходы', style='Regular.TButton',
            command=lambda: self.open_warehouse_history_window(consumption=True)
        ).pack(side=LEFT, padx=3, pady=3)

        ttk.Label(picture_subframe, image=self.image_rack_pil).pack(side=RIGHT, pady=1)

        get_info_log(user=user_data.get('user_name'), message='Warehouse mode widgets were rendered',
                     func_name=self.render_widgets_warehouse_mode.__name__, func_path=abspath(__file__))

    def render_widgets_selected_bom(self):
        """
        Функция рендера всех виджетов окна приложения в режиме просмотра BOM (спецификации) пресс-формы
        """
        define_table_name: Callable = lambda: f'BOM_HOT_RUNNER_{self.mold_number}' if self.hot_runner_bom else f'BOM_{self.mold_number}'
        part_number = self.get_value_by_selected_row(define_table_name(), 'NUMBER')
        # Рендер панели инструментов
        self.render_toolbar(back_func=self.get_molds_data,
                            add_row_func=lambda: self.render_typical_additional_window(
                                called_class=lambda: EditedBOM(self.mold_number, define_table_name()),
                                window_name='New Spare Part Information', access=user_data.get('molds_and_boms_data_changing'),
                                called_function=lambda: self.open_bom(
                                    self.mold_number)),
                            edit_row_func=self.render_bom_edition_window,
                            delete_row_func=lambda: self.delete_selected_table_row(f'BOM_{self.mold_number}', 'NUMBER'),
                            new_attachment_func=lambda: self.upload_attachment(bom_part=True),
                            looking_attachments_func=self.render_attachments_window)
        # Объявление основного и вложенных контейнеров для виджетов
        self.frame_main_widgets = Frame(self, relief=RIDGE)
        self.frame_main_widgets.pack(fill=X)
        main_sub_frame = Frame(self.frame_main_widgets)
        main_sub_frame.pack(fill=X, side=LEFT, pady=2, padx=2)
        picture_subframe = Frame(self.frame_main_widgets)
        picture_subframe.pack(side=RIGHT, pady=2, padx=2)
        title_frame = Frame(main_sub_frame)
        title_frame.pack(fill=BOTH, expand=True)
        description_frame = Frame(main_sub_frame)
        description_frame.pack(fill=BOTH, expand=True)
        bom_frame = ttk.LabelFrame(main_sub_frame, text='BOM', relief=RIDGE)
        bom_frame.pack(side=LEFT, padx=10, pady=1)
        sort_frame = ttk.LabelFrame(main_sub_frame, text='Сортировка по количеству', relief=RIDGE)
        sort_frame.pack(side=LEFT, padx=0, pady=1)
        stock_frame = ttk.LabelFrame(main_sub_frame, text='Склад', relief=RIDGE)
        stock_frame.pack(side=LEFT, padx=5, pady=1)
        # Рендер виджетов
        molds_data = table_funcs.TableInDb('All_molds_data', 'Database')
        mold_info = molds_data.get_table(type_returned_data='dict', first_param='MOLD_NUMBER', first_value=self.mold_number,
                                         last_string=True)
        (ttk.Label(title_frame, text=f'BOM для пресс-формы № {self.mold_number} ', style='Title.TLabel')
         .pack(side=LEFT, padx=8, pady=4))
        (ttk.Label(description_frame,
                   text=f'Проект: {mold_info.get("MOLD_NAME")} | Горячий канал: {mold_info.get("HOT_RUNNER_NUMBER")} | '
                        f'Год: {mold_info.get("RELEASE_YEAR")} | Кол-во гнёзд: {mold_info.get("CAVITIES_QUANTITY")}',
                   style='Regular.TLabel', font=('Arial', '9', 'normal'))
         .pack(side=LEFT, padx=8, pady=2))

        ttk.Button(
            bom_frame, text='Пресс-форма', style='Regular.TButton',
            command=lambda: self.open_bom(mold_number=self.mold_number)
        ).pack(side=LEFT, padx=6, pady=5)

        ttk.Button(
            bom_frame, text='Горячий канал', style='Regular.TButton',
            command=lambda: self.open_bom(mold_number=self.mold_number, hot_runner=True)
        ).pack(padx=6, pady=5, side=LEFT)

        ttk.Button(
            sort_frame, text='Все', style='Green.TButton',
            command=lambda: self.open_bom(mold_number=self.mold_number, hot_runner=self.hot_runner_bom)
        ).pack(padx=6, pady=5, side=LEFT)

        ttk.Button(
            sort_frame, text='В наличие', style='Yellow.TButton',
            command=lambda: self.open_bom(mold_number=self.mold_number, sort_status='В наличие',
                                          hot_runner=self.hot_runner_bom)
        ).pack(padx=6, pady=5, side=LEFT)

        ttk.Button(
            sort_frame, text='Отсутствующие', style='Coral.TButton',
            width=18,
            command=lambda: self.open_bom(mold_number=self.mold_number, sort_status='Отсутствующие',
                                          hot_runner=self.hot_runner_bom)
        ).pack(padx=6, pady=5, side=LEFT)

        ttk.Button(
            sort_frame, text='Меньше минимума', style='Gold.TButton',
            width=18,
            command=lambda: self.open_bom(mold_number=self.mold_number, sort_status='Меньше минимума',
                                          hot_runner=self.hot_runner_bom)
        ).pack(padx=6, pady=5, side=LEFT)

        ttk.Button(
            stock_frame, text='Списание', style='Regular.TButton',
            command=lambda: self.open_parts_quantity_changing_window(consumption=True)
        ).pack(padx=6, pady=5, side=LEFT)

        ttk.Button(
            stock_frame, text='Приход', style='Regular.TButton',
            command=self.open_parts_quantity_changing_window
        ).pack(padx=6, pady=5, side=LEFT)

        ttk.Label(picture_subframe, image=self.image_spare_part_pil).pack(side=RIGHT, pady=1)

        get_info_log(user=user_data.get('user_name'), message='BOM widgets were rendered',
                     func_name=self.render_widgets_selected_bom.__name__, func_path=abspath(__file__))

    def on_clicked_or_pressed_table_row(self, event):
        """
        Обработчик события двойного нажатия мыши или клавиши Enter на выделеную строку таблицы
        :param event: Обрабатываемое событие
        """
        self.open_bom()
        self.event = event

    def on_pressed_key_one(self, event):
        """
        Обработчик события нажатия клавиши 1
        :param event: Обрабатываемое событие
        """
        self.open_main_menu()

    def on_pressed_key_two(self, event):
        """
        Обработчик события нажатия клавиши 2
        :param event: Обрабатываемое событие
        """
        try:
            self.key_two_func()
        except TypeError:
            pass

    def on_pressed_key_three(self, event):
        """
        Обработчик события нажатия клавиши 3
        :param event: Обрабатываемое событие
        """
        try:
            self.key_three_func()
        except TypeError:
            pass

    def on_pressed_key_four(self, event):
        """
        Обработчик события нажатия клавиши 4
        :param event: Обрабатываемое событие
        """
        try:
            self.key_four_func()
        except TypeError:
            pass

    def on_pressed_key_five(self, event):
        """
        Обработчик события нажатия клавиши 5
        :param event: Обрабаьываемое событие
        """
        try:
            self.key_five_func()
        except TypeError:
            pass

    def render_table(self, columns_sizes: dict):
        """
        Вывод таблицы в отображаемом окне приложения на основе полученных данных, которые были предварительно
        загруженны из таблицы базы данных
        :param columns_sizes: Информация о размере каждого столбца таблицы
        """
        # определяем столбцы
        columns = self.current_table[0]
        # определяем таблицу
        self.tree = ttk.Treeview(columns=columns, show="headings")
        self.tree.pack(fill=BOTH, expand=1)
        # определяем заголовки
        for col_name in columns:
            self.tree.heading(col_name, text=col_name)
        # настраиваем столбцы
        for col_num, col_size in columns_sizes.items():
            self.tree.column(column=col_num, stretch=YES, width=col_size)
        # Добавление данных в отображаемую таблицу. Чтобы последняя добавленная информацию отображалась необходимо
        # прежде сделать реверсию массива данных
        reversed_data = list(reversed(self.current_table[1:]))
        for row in reversed_data:
            self.tree.insert("", END, values=row)

        get_info_log(user=user_data.get('user_name'), message='Table was rendered',
                     func_name=self.render_table.__name__, func_path=abspath(__file__))
        # # добавляем вертикальную прокрутку
        # scrollbar = ttk.Scrollbar(self.tree, orient=VERTICAL, command=self.tree.yview)
        # self.tree.configure(yscrollcommand=scrollbar.set)
        # # scrollbar.grid(row=0, column=1, sticky="ns")
        # scrollbar.pack(side=RIGHT)

    def get_row_number_in_table(self) -> int:
        """
        Данная функция преобразует полученный ID номер выделенной строки пользователем в таблице  в порядковый номер строки.
        В библиотеке Tkinter ID присваивается следующим образом: I001, I002, ..., I009, I00A, I00B, I00C, I00D, I00E, I00F,
        I010, I011, ..., I01A, I01B, I01C, I01D, I01E, I01F, I020, I021, ....
        Соответственно для корректного получения порядкового номера строки необходимо учесть числовые и буквенные итерации.
        :return: порядковый номер выделенной строки
        """

        tree_row_id = self.tree.focus().replace('I', '')
        if tree_row_id:
            cycle_number = int(tree_row_id[:-1])

            try:
                row_id = int(tree_row_id) - 1
            except ValueError:
                letter_values = {'A': 0, 'B': 1, 'C': 2, 'D': 3, 'E': 4, 'F': 5}
                indicated_letter = tree_row_id[-1]
                cycle_number = int(tree_row_id[:-1])
                row_id = int(f'{cycle_number + 1}{letter_values.get(indicated_letter)}') - 1

            return row_id + cycle_number * 6

    def get_molds_data(self, sort_status: str | bool = None):
        """
        Функция вывода перечня всех пресс-форм в табличном виде в окне приложения
        :param sort_status: Текстовое значение указывающее на необходимость проведения сортировки перечня п/ф
            по определённому критерию ("IN"; "OUT"; "IN SERVICE").
        """
        # Выгрузка из базы данных перечня пресс-форм, сортировка и запись в переменные класса
        self.sort_molds()
        # Очистка области в окне приложения перед выводом новой таблицы
        self.clean_frames()
        if sort_status:
            self.sort_status = sort_status
            self.current_table = self.sorted_molds_data_tuple.get(sort_status)
        else:
            self.sort_status = None
        # Рендер кнопок для данного окна
        self.render_widgets_molds_list()
        # Определение размера столбцов таблицы
        columns_sizes = {'#1': 5, '#2': 10, '#3': 35, '#4': 25, '#5': 10, '#6': 10, '#7': 20, '#8': 20, '#9': 20,
                         '#10': 20}
        # Рендер таблицы
        self.render_table(columns_sizes=columns_sizes)
        # Объявление обработчиков событий на двойной клик мышью или клавиши Enter
        self.tree.bind('<Double-ButtonPress-1>', self.on_clicked_or_pressed_table_row)
        self.tree.bind('<Return>', self.on_clicked_or_pressed_table_row)
        # Объявление других обработчиков событий
        self.add_listeners(funk_two=self.open_main_menu,
                           funk_three=lambda: self.render_typical_additional_window(
                               called_class=EditedMold, window_name='New Mold Information',
                               called_function=self.get_molds_data),
                           funk_four=self.render_mold_edition_window)

    def add_listeners(self, funk_two: Callable, funk_three: Callable = None,
                      funk_four: Callable = None, funk_five: Callable = None):
        """
        Функция добавления обработчиков событий при нажатии клавиш клавиатуры
        :param funk_two: Вызываемая функция клавиши 2
        :param funk_three: Вызываемая функция клавиши 3
        :param funk_four: Вызываемая функция клавиши 4
        :param funk_five: Вызываемая функция клавиши 5
        """
        self.key_two_func = funk_two
        self.key_three_func = funk_three
        self.key_four_func = funk_four
        self.key_five_func = funk_five
        self.master.bind('1', self.on_pressed_key_one)
        self.master.bind('2', self.on_pressed_key_two)
        self.master.bind('3', self.on_pressed_key_three)
        self.master.bind('4', self.on_pressed_key_four)
        self.master.bind('5', self.on_pressed_key_five)

    def remove_listeners(self):
        """
        Функция удаления удаления обработчиков событий
        """
        for key in range(1, 6):
            self.master.unbind(f'{key}')

    def sort_molds(self):
        """
        Функция сортировки общего перечня пресс-форм в зависимости от её статуса и записи информации в переменные класса.
        Вызывается при открытие окна с перечнем п/ф. Функция необходима для минимизации количества вызовов базы данных.
        """
        # Перезапись переменных для обновления данных, которые будут получены из БД
        self.sorted_molds_data_tuple = {status: [] for status in mold_statuses_list}
        self.sorted_molds_data_dict = {status: [] for status in mold_statuses_list}
        # Выгрузка таблицы из БД
        molds_data = table_funcs.TableInDb('All_molds_data', 'Database')
        self.all_molds_table_dict = molds_data.get_table(type_returned_data='dict')
        self.molds_list_data = molds_data.get_table(type_returned_data='tuple')
        self.current_table = molds_data.get_table(type_returned_data='tuple')
        # Сортировка и запись в зависимости от статуса
        for num, mold_info in enumerate(self.all_molds_table_dict):
            if num == 0:
                for status in mold_statuses_list:
                    self.sorted_molds_data_dict.get(status).append(mold_info)
                    self.sorted_molds_data_tuple.get(status).append(self.molds_list_data[num])
            else:
                status = mold_info.get('STATUS')
                self.sorted_molds_data_dict.get(status).append(mold_info)
                self.sorted_molds_data_tuple.get(status).append(self.molds_list_data[num])

    def sort_bom_parts(self, hot_runner: bool = None):
        """
        Функция сортировки перечня BOM в зависимости от имеющегося количества запчастей на складе.
        Вызывается при открытие окна с BOM. Функция необходима для минимизации количества вызовов базы данных.
        :param hot_runner: Булево значение, которое характеризует какой тип BOM был выбран (Пресс-форма или горячий канал)
        """
        # Перезапись переменных для обновления данных, которые будут получены из БД
        self.sorted_bom_tuple = {status: [] for status in part_statuses_list}
        self.sorted_bom_dict = {status: [] for status in part_statuses_list}
        # Выгрузка таблицы из БД
        define_table_name: Callable = lambda: f'BOM_HOT_RUNNER_{self.mold_number}' if hot_runner else f'BOM_{self.mold_number}'
        bom = table_funcs.TableInDb(define_table_name(), 'Database')
        self.bom_table_dict = bom.get_table(type_returned_data='dict')
        self.bom_data = bom.get_table(type_returned_data='tuple')
        self.current_table = bom.get_table(type_returned_data='tuple')
        # Сортировка и запись в зависимости от статуса
        for num, part_info in enumerate(self.bom_table_dict):
            if num == 0:
                for status in part_statuses_list:
                    self.sorted_bom_dict.get(status).append(part_info)
                    self.sorted_bom_tuple.get(status).append(self.bom_data[num])
            else:
                # Проверка числовых значений и сортировка данных по условиям
                try:
                    min_percent = int(part_info.get('MIN_PERCENT'))
                    parts_qantity = int(part_info.get('PARTS_QUANTITY'))
                    pcs_in_mold = int(part_info.get('PCS_IN_MOLDS'))
                    current_percent = int(parts_qantity) / int(pcs_in_mold) * 100
                except TypeError:
                    pass
                else:

                    if parts_qantity > 0:
                        self.sorted_bom_dict.get('В наличие').append(part_info)
                        self.sorted_bom_tuple.get('В наличие').append(self.bom_data[num])
                        if current_percent < min_percent:
                            self.sorted_bom_dict.get('Меньше минимума').append(part_info)
                            self.sorted_bom_tuple.get('Меньше минимума').append(self.bom_data[num])
                    elif parts_qantity == 0:
                        self.sorted_bom_dict.get('Отсутствующие').append(part_info)
                        self.sorted_bom_tuple.get('Отсутствующие').append(self.bom_data[num])

    def get_value_by_selected_row(self, table_name: str, column_name: str) -> str:
        """
        Получение определенного значения (например: номер пресс-формы или номер элемента в BOM) из 
        выделенной строки таблицы пользователем (table_row_number) в окне приложения
        :param column_name: Имя столбца / параметра по которому будет искаться строка
        :param table_name: Имя таблицы из базы данных
        """
        # Получения номера выделенной строки
        table_row_number = self.get_row_number_in_table()
        # Получение таблицы
        if self.sort_status:
            table = self.sorted_molds_data_dict.get(self.sort_status)
        else:
            db = table_funcs.TableInDb(table_name, 'Database')
            table = db.get_table(type_returned_data='dict')
        # Так как выведенная таблица в окне приложения перевернута, необходимо сделать реверсию полученного массива
        reversed_table = list(reversed(table))
        # Получение значения из массива данных
        try:
            return reversed_table[table_row_number].get(column_name)
        except (TypeError, IndexError):
            pass

    def open_bom(self, mold_number: str = None, hot_runner: bool = None, sort_status: str | bool = None):
        """
        Функция для вывода спецификации (BOM) пресс-формы в табличном виде в окне приложения
        :param sort_status: Булево значение, которое характеризует был ли выбран параметр сортировки для отображения таблицы
        :param hot_runner: Булево значение, которое характеризует какой тип BOM был выбран (Пресс-форма или горячий канал)
        :param mold_number: Номер пресс-формы полученный из строки ввода
        """
        self.mold_number_entry_field.delete(0, END)
        self.hot_runner_bom = hot_runner
        define_table_name: Callable = lambda: f'BOM_HOT_RUNNER_{mold_number}' if hot_runner else f'BOM_{mold_number}'
        if not mold_number:
            mold_number = self.get_value_by_selected_row('All_molds_data', 'MOLD_NUMBER')
        try:
            # Выгрузка информации из базы данных
            if sort_status:
                self.sort_status = sort_status
                self.sort_bom_parts(hot_runner)
                self.current_table = self.sorted_bom_tuple.get(sort_status)
            else:
                self.sort_status = None
                self.mold_number = mold_number
                bom = table_funcs.TableInDb(define_table_name(), 'Database')
                self.current_table = bom.get_table(type_returned_data='tuple')
        except sqlite3.OperationalError:
            messagebox.showerror('Уведомление об ошибке', f'Спецификации по номеру "{mold_number}" не имеется')
        else:
            # Очистка области в окне приложения перед выводом новой таблицы
            self.clean_frames()
            # Обновление обработчиков событий
            self.remove_listeners()
            self.add_listeners(funk_two=self.get_molds_data,
                               funk_three=lambda: self.render_typical_additional_window(
                                   called_class=lambda: EditedBOM(self.mold_number),
                                   window_name='New Spare Part Information',
                                   called_function=lambda: self.open_bom(
                                       self.mold_number)),
                               funk_four=self.render_bom_edition_window)
            self.render_widgets_selected_bom()
            # Определение размера столбцов таблицы
            columns_sizes = {'#1': 5, '#2': 10, '#3': 35, '#4': 25, '#5': 10, '#6': 10, '#7': 20, '#8': 20, '#9': 20,
                             '#10': 20, '#11': 20, '#12': 20, }
            self.render_table(columns_sizes=columns_sizes)

    def open_mold_scanning_window(self):
        """
        Функция вывода окна с виджетами для смены статуса пресс-формы и таблицей с историей этих изменений
        """
        # Очистка области в окне приложения перед выводом новой таблицы
        self.clean_frames()
        # Формирование табличных данных
        moving_history = table_funcs.TableInDb('Molds_moving_history', 'Database')
        moving_history_data = moving_history.get_table(type_returned_data='tuple')
        self.current_table = []
        self.current_table = [columns_molds_moving_history_table if i == 0 else moving_history_data[i - 1]
                              for i in range(0, len(moving_history_data) + 1)]
        # Рендер виджетов
        self.render_widgets_mold_scanning_mode()
        # Определение размера столбцов таблицы
        self.render_table(columns_sizes=columns_sizes_moving_history_table)
        # Объявление обработчиков событий
        self.add_listeners(funk_two=self.open_main_menu)

    def open_warehouse_history_window(self, consumption: bool = None):
        """
        Функция вывода окна с виджетами для смены статусов пресс-форм и таблицей с историей их изменений
        :param consumption: Булево значение, которое принимает значение True когда происходит расход
        """
        define_table_name: Callable = lambda: 'OUT_warehouse_history' if consumption else 'IN_warehouse_history'
        # Очистка области в окне приложения перед выводом новой таблицы
        self.clean_frames()
        # Формирование табличных данных
        warehouse_history = table_funcs.TableInDb(define_table_name(), 'Database')
        warehouse_history_data = warehouse_history.get_table(type_returned_data='tuple')
        self.current_table = []
        self.current_table = [columns_warehouse_table if i == 0 else warehouse_history_data[i - 1]
                              for i in range(0, len(warehouse_history_data) + 1)]
        # Рендер виджетов
        self.render_widgets_warehouse_mode(consumption)
        # Определение размера столбцов таблицы
        self.render_table(columns_sizes=columns_sizes_warehouse_table)
        # Объявление обработчиков событий
        self.add_listeners(funk_two=self.open_main_menu)

    def open_qr_window(self, next_status: str):
        """
        Рендер окна для получения сообщения от сканера QR кода
        :param next_status: Новое значение параметра STATUS из таблицы перечня всех пресс-форм, которое заменит старое 
        """
        if user_data.get('mold_status_changing'):
            qr_window = QRWindow(next_status)
            qr_window.title('QR code getting')
            qr_window.resizable(False, False)
            qr_window.render_widgets_before_getting_code()
            if qr_window.changed_data:
                get_info_log(user=user_data.get('user_name'), message='Status changing',
                             func_name=self.open_qr_window.__name__, func_path=abspath(__file__))
                self.tree.pack_forget()
                self.open_mold_scanning_window()

    def render_typical_additional_window(self, called_class: Callable, window_name: str, access: str = None,
                                         called_function: Callable = None):
        """
        Функция создания дополнительного окна по шаблону
        :param access: Переменная характеризующая наличие доступа у пользователя для выбранного действия
        :param called_function: Вызываемая функция в случае изменения каких либо данных после взаимодействия
        пользователя в открытом окне
        :param window_name: Название открываемого окна
        :param called_class: Вызываемый класс для создания его экземпляра
        """
        if access != 'False':
            window = called_class()
            window.title(window_name)
            window.resizable(False, False)
            window.render_widgets()
            if window.changed_data:
                get_info_log(user=user_data.get('user_name'), message='Successful data changing',
                             func_name=self.render_typical_additional_window.__name__, func_path=abspath(__file__))
                self.tree.pack_forget()
                if called_function:
                    called_function()
        else:
            messagebox.showerror('Ошибка доступа',
                                 'У Вас нет доступа. Для его предоставления обратитесь к администратору')

    def render_upload_bom_window(self):
        """
        Рендер окна для выбора типа загружаемого BOM (пресс-форма или горячий канал)
        """
        window = tkinter.Toplevel()
        window.title('New BOM uploading')
        window.resizable(False, False)
        window.focus_set()
        window.grab_set()
        (ttk.Label(window, text='Выберите тип загружаемого BOM', font=('Times', '11', 'normal'))
         .pack(side=TOP, pady=5))
        ttk.Button(
            window, text='Пресс-форма', style='Regular.TButton',
            width=20,
            command=lambda: self.open_file_and_download(window)
        ).pack(side=LEFT, padx=5, pady=5)
        ttk.Button(
            window, text='Горячий канал', style='Regular.TButton',
            width=20,
            command=lambda: self.open_file_and_download(window, hot_runner=True)
        ).pack(side=LEFT, padx=5, pady=5)
        window.mainloop()

    def render_mold_edition_window(self):
        """
        Рендер окна для редактирования данных о пресс-форме
        """
        mold_data_edition_access = user_data.get('molds_and_boms_data_changing')
        mold_number = self.mold_number_entry_field.get()
        if not mold_number:
            mold_number = self.get_value_by_selected_row('All_molds_data', 'MOLD_NUMBER')
        # Если получен номер елемента таблицы, тогда будет вызвано окно для взаимодействия с пользователем
        if mold_number:

            try:
                # Выгрузка информации о пресс-форме из базы данных
                all_molds_data = table_funcs.TableInDb('All_molds_data', 'Database')
                mold_data = all_molds_data.get_table(type_returned_data='dict', first_param='MOLD_NUMBER',
                                                     first_value=mold_number, last_string=True)
            except (sqlite3.OperationalError, IndexError):
                messagebox.showerror('Уведомление об ошибке', f'Данных о пресс-форме "{mold_number}" не имеется')
                get_warning_log(user=user_data.get('user_name'), message='sqlite3.OperationalError',
                                func_name=self.render_mold_edition_window.__name__, func_path=abspath(__file__))
            else:
                self.render_typical_additional_window(called_class=lambda: EditedMold(mold_data=mold_data),
                                                      window_name='Mold Data Edition', access=mold_data_edition_access,
                                                      called_function=self.get_molds_data)

    def render_bom_edition_window(self):
        """
        Рендер окна для редактирования данных запчастях открытого в приложении BOM
        """
        bom_edition_access = user_data.get('molds_and_boms_data_changing')
        define_table_name: Callable = lambda: f'BOM_HOT_RUNNER_{self.mold_number}' if self.hot_runner_bom else f'BOM_{self.mold_number}'
        table_name = define_table_name()
        part_number = self.get_value_by_selected_row(table_name, 'NUMBER')
        # Если получен номер елемента таблицы, тогда будет вызвано окно для взаимодействия с пользователем
        if part_number:

            try:
                # Выгрузка информации о пресс-форме из базы данных
                bom = table_funcs.TableInDb(table_name, 'Database')
                part_data = bom.get_table(type_returned_data='dict', first_param='NUMBER',
                                          first_value=part_number, last_string=True)
            except (sqlite3.OperationalError, IndexError):
                messagebox.showerror('Уведомление об ошибке', f'Данных о запчасти не имеется')
                get_warning_log(user=user_data.get('user_name'), message='sqlite3.OperationalError',
                                func_name=self.render_bom_edition_window.__name__, func_path=abspath(__file__))
            else:
                self.render_typical_additional_window(
                    called_class=lambda: EditedBOM(part_data=part_data, mold_number=self.mold_number,
                                                   table_name=table_name),
                    window_name='BOM Edition', access=bom_edition_access,
                    called_function=lambda: self.open_bom(self.mold_number))

    def render_attachments_window(self):
        """
        Рендер окна для просмотра прикреплённых вложенных файлов к номеру пресс-формы, либо номеру запчасти
        """
        define_table_name: Callable = lambda: f'BOM_HOT_RUNNER_{self.mold_number}' if self.hot_runner_bom else f'BOM_{self.mold_number}'
        table_name = define_table_name()
        part_number = self.get_value_by_selected_row(table_name, 'NUMBER')
        # Если получен номер елемента таблицы, тогда будет вызвано окно для взаимодействия с пользователем
        if part_number:
            self.render_typical_additional_window(called_class=lambda: Attachment(mold_number=self.mold_number,
                                                                                  part_number=part_number,
                                                                                  hot_runner=self.hot_runner_bom),
                                                  window_name='Attachments')

    def open_parts_quantity_changing_window(self, consumption: bool = None):
        """
        Вывод окна для взаимодействия со складом. Взятие каких либо запчастей, либо её приём на склад. 
        :param consumption: Булево значение, которое принимает значение True когда происходит расход
        """
        stock_changing_access = user_data.get('stock_changing')
        define_table_name: Callable = lambda: f'BOM_HOT_RUNNER_{self.mold_number}' if self.hot_runner_bom else f'BOM_{self.mold_number}'
        table_name = define_table_name()
        part_number = self.get_value_by_selected_row(table_name, 'NUMBER')
        # Если получен номер елемента таблицы, тогда будет вызвано окно для взаимодействия с пользователем
        if part_number:

            try:
                # Выгрузка информации о пресс-форме из базы данных
                bom = table_funcs.TableInDb(table_name, 'Database')
                part_data = bom.get_table(type_returned_data='dict', first_param='NUMBER',
                                          first_value=part_number, last_string=True)
            except (sqlite3.OperationalError, IndexError):
                messagebox.showerror('Уведомление об ошибке', f'Данных о запчасти не имеется')
                get_warning_log(user=user_data.get('user_name'), message='sqlite3.OperationalError',
                                func_name=self.render_bom_edition_window.__name__, func_path=abspath(__file__))
            else:
                self.render_typical_additional_window(
                    called_class=lambda: Stock(part_data=part_data, mold_number=self.mold_number,
                                               consumption=consumption, table_name=table_name),
                    window_name='BOM Edition', access=stock_changing_access,
                    called_function=lambda: self.open_bom(self.mold_number))

    def delete_selected_table_row(self, table_name: str, column_name: str):
        """
        Фнкция удаления строки из таблицы базы данных на основании выделенной строки
        :param table_name: Имя таблицы из базы данных
        :param: column_name: Имя столбца / параметра по которому будет искаться строка для удаления
        """
        if user_data.get('molds_and_boms_data_changing'):                
            message = "Вы уверены, что хотите удалить данную строку"
            if messagebox.askyesno(title='Подтверждение', message=message, parent=self):
    
                try:
                    number = self.get_value_by_selected_row(table_name, column_name)
                    db = table_funcs.TableInDb(table_name, 'Database')
                    db.delete_data(column_name=column_name, value=number)
                except sqlite3.OperationalError:
                    messagebox.showerror('Уведомление об ошибке', 'Ошибка удаления. Обратитесь к администратору.')
                    get_info_log(user=user_data.get('user_name'), message='sqlite3.OperationalError',
                                 func_name=self.delete_selected_table_row.__name__, func_path=abspath(__file__))
                else:
                    messagebox.showinfo('Уведомление', 'Удаление успешно произведено!')
                    get_info_log(user=user_data.get('user_name'), message='Row was deleted from table',
                                 func_name=self.delete_selected_table_row.__name__, func_path=abspath(__file__))
    
                    if column_name == 'MOLD_NUMBER':
    
                        try:
                            bom = table_funcs.TableInDb(f'BOM_{number}', 'Database')
                            bom.delete_db_table()
                        except sqlite3.OperationalError:
                            pass
                        self.tree.pack_forget()
                        self.get_molds_data()
                    else:
                        self.tree.pack_forget()
                        self.open_bom(self.mold_number)

    def delete_selected_bom(self):
        """
        Фнкция удаления BOM таблицы из базы данных на основании выделенной строки
        """
        if user_data.get('molds_and_boms_data_changing'):
            mold_number = self.mold_number_entry_field.get()
            if not mold_number:
                mold_number = self.get_value_by_selected_row('All_molds_data', 'MOLD_NUMBER')
    
            if mold_number:
                message = "Вы уверены, что хотите удалить данный BOM"
                if messagebox.askyesno(title='Подтверждение', message=message, parent=self):
    
                    try:
                        bom = table_funcs.TableInDb(f'BOM_{mold_number}', 'Database')
                        bom.delete_db_table()
                        try:
                            hot_runner_bom = table_funcs.TableInDb(f'BOM_HOT_RUNNER_{mold_number}', 'Database')
                            hot_runner_bom.delete_db_table()
                        except sqlite3.OperationalError:
                            pass
                    except sqlite3.OperationalError:
                        messagebox.showerror('Уведомление об ошибке',
                                             'Ошибка удаления. Вероятнее всего данный BOM отсутствовал.')
                        get_warning_log(user=user_data.get('user_name'), message='sqlite3.OperationalError',
                                        func_name=self.delete_selected_table_row.__name__, func_path=abspath(__file__))
                    else:
                        messagebox.showinfo('Уведомление', 'Удаление BOM успешно произведено!')
                        get_info_log(user=user_data.get('user_name'), message='Row was deleted from table',
                                     func_name=self.delete_selected_table_row.__name__, func_path=abspath(__file__))

    def confirm_delete(self):
        """
        Фнкция вывода диалогового окна для запроса подтверждения закрытия окна
        """
        message = "Вы уверены, что хотите удалить данную строку"
        if messagebox.askyesno(title='Подтверждение', message=message, parent=self):
            return True


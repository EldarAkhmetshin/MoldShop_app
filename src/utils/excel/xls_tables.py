import os
from os.path import abspath
from tkinter import messagebox, filedialog
from openpyxl import load_workbook
import openpyxl
from openpyxl.worksheet.dimensions import ColumnDimension, DimensionHolder
from openpyxl.utils import get_column_letter
from pandas import DataFrame

from src.data import purchased_statuses
from src.global_values import user_data
from src.utils.logger.logs import get_warning_log
from src.utils.sql_database import table_funcs


def get_new_bom_from_excel_file(file_path: str, work_sheet_name: str) -> (tuple, list, bool, str):
    """
    Функция для получения информации из Excel файла
    :param file_path: Путь к Excel файлу
    :param work_sheet_name: Наименование листа Excel файла откуда будет браться информация
    :return: Кортеж из наименований столбцов таблицы
    :return: Массив данных из полученной таблицы. Каждая строка таблицы записана в отдельный кортеж
    :return: Булево значение, характеризующее статус получения информации из файла
    (True если получены корректные данные)
    :return: Текст ошибки
    """
    column_names = None
    rows_data = []

    work_book = load_workbook(file_path)
    try:
        work_sheet = work_book[f'{work_sheet_name}']
    except KeyError:
        messagebox.showerror('Уведомление об ошибке',
                             'Название листа не соответствует. Убедитесь, что вы используете '
                             'правильный файл шаблона')
    else:
        # part_nums = []
        # for count, row in enumerate(work_sheet.values):
        #     # Проверка на наличие номера запчасти в каждой строке и на то, чтобы этот номер не повторялся в BOM
        #     part_num = row[0]
        #     pcs_in_mold = row[2]
        #     new_parts_in_stock = row[8]
        #     old_parts_in_stock = row[9]
        #     min_percent = row[11]
        #
        #     if not part_num:
        #         return (column_names, rows_data,
        #                 False, 'BOM не может быть загружен, так как имеется строка без номера запчасти')
        #     elif part_num in part_nums:
        #         return (column_names, rows_data, False,
        #                 f'BOM не может быть загружен, так как номер запчасти: {part_num} '
        #                 f'дублируется в строке {count + 1}')
        #     elif count != 0 and not isinstance(pcs_in_mold, int):
        #         return (column_names, rows_data, False,
        #                 f'BOM не может быть загружен, так как значение "Кол-во в пресс-форме, шт": {pcs_in_mold}'
        #                 f' должно быть числом в строке {count + 1}')
        #     elif count != 0 and not isinstance(new_parts_in_stock, int):
        #         return (column_names, rows_data, False,
        #                 f'BOM не может быть загружен, так как значение "Кол-во в пресс-форме, шт": {new_parts_in_stock}'
        #                 f' должно быть числом в строке {count + 1}')
        #     elif count != 0 and not isinstance(old_parts_in_stock, int):
        #         return (column_names, rows_data, False,
        #                 f'BOM не может быть загружен, так как значение "Кол-во в пресс-форме, шт": {old_parts_in_stock}'
        #                 f' должно быть числом в строке {count + 1}')
        #     elif count != 0 and not isinstance(min_percent, int):
        #         return (column_names, rows_data, False,
        #                 f'BOM не может быть загружен, так как значение "Кол-во в пресс-форме, шт": {min_percent}'
        #                 f' должно быть числом в строке {count + 1}')
        #
        #     if count == 0:
        #         column_names = row
        #     else:
        #         rows_data.append(row)
        #     part_nums.append(part_num)
        for count, row in enumerate(work_sheet.values):
            if count == 0:
                column_names = row
            else:
                rows_data.append(row)

        print(44, column_names, rows_data)
        return column_names, rows_data, True, None


def get_purchasing_list_from_excel_file(file_path: str, work_sheet_name: str) -> (tuple, list, bool, str):
    """
    Функция для получения информации из Excel файла
    :param file_path: Путь к Excel файлу
    :param work_sheet_name: Наименование листа Excel файла откуда будет браться информация
    :return: Кортеж из наименований столбцов таблицы
    :return: Массив данных из полученной таблицы. Каждая строка таблицы записана в отдельный кортеж
    :return: Булево значение, характеризующее статус получения информации из файла
    (True если получены корректные данные)
    :return: Текст ошибки
    """
    column_names = None
    rows_data = []

    work_book = load_workbook(file_path)
    try:
        work_sheet = work_book[f'{work_sheet_name}']
    except KeyError:
        messagebox.showerror('Уведомление об ошибке',
                             'Название листа не соответствует. Убедитесь, что вы используете '
                             'правильный файл шаблона')
    else:
        part_nums = []
        for count, row in enumerate(work_sheet.values):
            # Проверка на наличие номера запчасти в каждой строке и на то, чтобы этот номер не повторялся в BOM
            part_num = row[0]
            if not part_num:
                return (column_names, rows_data,
                        False, 'Информация не может быть загружена, так как имеется строка без номера запчасти')

            if count == 0:
                column_names = row
            else:
                rows_data.append(row)
            part_nums.append(part_num)

        return column_names, rows_data, True, None


def export_excel_table(table: list):
    """
    Функция сохранения открытой таблицы из окна приложения в Иксель файл формата xlsx
    """
    # Открытие диалогового окна для сохранения файла пользователем в локальной директории компьютера,
    # с дальнейшим извлечением пути к выбранному файлу в виде строки
    file_path = filedialog.asksaveasfilename(
        filetypes=(('XLSX files', '*.xlsx'),)
    )
    if file_path:
        file_path = file_path.replace('.xlsx', '')
        # Формирование Иксель файла типа xlsx
        table_length = len(table)
        saved_table = {}
        for num, column_name in enumerate(table[0]):
            saved_table[column_name] = tuple(table[i][num] for i in range(1, table_length))
        try:
            df = DataFrame(saved_table)
            df.to_excel(excel_writer=f'{file_path}.xlsx', sheet_name='Table', index=False)
            wb = openpyxl.load_workbook(f'{file_path}.xlsx')
            ws = wb['Table']
            dim_holder = DimensionHolder(worksheet=ws)

            for col in range(ws.min_column, ws.max_column + 1):
                dim_holder[get_column_letter(col)] = ColumnDimension(ws, min=col, max=col, width=20)

            ws.column_dimensions = dim_holder
            wb.save(f'{file_path}.xlsx')
        except Exception:
            messagebox.showerror(title='Уведомление об ошибке',
                                 message='Ошибка в записи файла.\nПовторите ещё раз, либо обратитесь к администратору')
            get_warning_log(user=user_data.get('user_name'), message='Table wasnt saved in Excel file',
                            func_name=export_excel_table.__name__, func_path=abspath(__file__))
        else:
            messagebox.showinfo(title='Уведомление', message='Таблица успешно сохранена на Ваш компьютер')


def save_new_molds_list():
    # Получение данных из Иксель файла
    column_names, rows_data = get_new_bom_from_excel_file(
        file_path=os.path.abspath(os.path.join('..', 'incoming_data', 'molds_data.xlsx')),
        work_sheet_name='Molds')
    # Загрузка данных в базу данных
    molds_data = table_funcs.TableInDb('All_molds_data', 'Database')
    for row in rows_data:
        molds_data.insert_data(row)
